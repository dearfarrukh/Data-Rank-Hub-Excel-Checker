
import os
import re
import io
import hashlib
from difflib import get_close_matches
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# =========================================================
# DATA RANK HUB — SIMPLE EXCEL CHECKER
# Upload → Check → Fix → Download
# Same app can run locally or from Google Drive / Colab.
# =========================================================

st.set_page_config(
    page_title="Data Rank Hub Excel Checker",
    layout="wide"
)

st.title("Data Rank Hub Excel Checker")
st.caption("Upload → Check → Fix → Download")

APP_FOLDER = os.path.dirname(os.path.abspath(__file__))

# Keep flag folders relative to the app so the same code works
# on the home PC and inside a Google Drive / Colab app folder.
COUNTRY_FLAG_LIBRARY = os.path.join(
    APP_FOLDER, "flags", "flag_library"
)
STATE_FLAG_LIBRARY = os.path.join(
    APP_FOLDER, "states"
)

IMAGE_EXTENSIONS = (".png", ".jpg", ".jpeg")

ENTITY_CANDIDATES = [
    "Country", "State", "Entity", "City", "Company", "Name"
]

MONTH_NAMES = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12
}


# =========================================================
# HELPERS
# =========================================================

def detect_entity_column(df):
    for candidate in ENTITY_CANDIDATES:
        if candidate in df.columns:
            return candidate
    return df.columns[0]


def normalize_name(text):
    text = str(text).strip().lower()
    replacements = {
        "&": "and",
        "_": " ",
        "-": " ",
        ".": "",
        "'": "",
        "’": ""
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return " ".join(text.split())


def parse_period_column(column):
    text = str(column).strip()

    # Annual
    if re.fullmatch(r"\d{4}", text):
        year = int(text)
        if 1 <= year <= 2200:
            return {
                "type": "Annual",
                "year": year,
                "month": 1,
                "quarter": 1,
                "sort_key": year,
                "original_column": column,
                "display": text
            }

    # Monthly MM-YYYY / MM / YYYY
    match = re.fullmatch(r"(\d{1,2})\s*[-/]\s*(\d{4})", text)
    if match:
        month = int(match.group(1))
        year = int(match.group(2))
        if 1 <= month <= 12:
            return {
                "type": "Monthly",
                "year": year,
                "month": month,
                "quarter": ((month - 1) // 3) + 1,
                "sort_key": year * 12 + month,
                "original_column": column,
                "display": text
            }

    # Monthly YYYY-MM
    match = re.fullmatch(r"(\d{4})\s*[-/]\s*(\d{1,2})", text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return {
                "type": "Monthly",
                "year": year,
                "month": month,
                "quarter": ((month - 1) // 3) + 1,
                "sort_key": year * 12 + month,
                "original_column": column,
                "display": text
            }

    # Monthly Jan 2024
    match = re.fullmatch(r"([A-Za-z]+)\s+(\d{4})", text)
    if match:
        month_name = match.group(1).lower()
        year = int(match.group(2))
        if month_name in MONTH_NAMES:
            month = MONTH_NAMES[month_name]
            return {
                "type": "Monthly",
                "year": year,
                "month": month,
                "quarter": ((month - 1) // 3) + 1,
                "sort_key": year * 12 + month,
                "original_column": column,
                "display": text
            }

    # Quarterly 2024 Q1
    match = re.fullmatch(r"(\d{4})\s*[- ]?Q([1-4])", text, re.IGNORECASE)
    if match:
        year = int(match.group(1))
        quarter = int(match.group(2))
        return {
            "type": "Quarterly",
            "year": year,
            "month": (quarter - 1) * 3 + 1,
            "quarter": quarter,
            "sort_key": year * 4 + quarter,
            "original_column": column,
            "display": text
        }

    # Quarterly Q1 2024
    match = re.fullmatch(r"Q([1-4])\s*[- ]?(\d{4})", text, re.IGNORECASE)
    if match:
        quarter = int(match.group(1))
        year = int(match.group(2))
        return {
            "type": "Quarterly",
            "year": year,
            "month": (quarter - 1) * 3 + 1,
            "quarter": quarter,
            "sort_key": year * 4 + quarter,
            "original_column": column,
            "display": text
        }

    return None


def detect_period_columns(df):
    detected = []
    for column in df.columns:
        parsed = parse_period_column(column)
        if parsed is not None:
            detected.append(parsed)

    if not detected:
        return [], None

    counts = {"Annual": 0, "Monthly": 0, "Quarterly": 0}
    for item in detected:
        counts[item["type"]] += 1

    dominant_type = max(counts, key=counts.get)
    filtered = [
        item for item in detected
        if item["type"] == dominant_type
    ]
    return filtered, dominant_type


def is_true_blank(value):
    if value is None:
        return True
    if pd.isna(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def expected_period_keys(periods, time_type):
    if not periods:
        return []

    sorted_periods = sorted(periods, key=lambda x: x["sort_key"])
    first = sorted_periods[0]
    last = sorted_periods[-1]
    expected = []

    if time_type == "Annual":
        for year in range(first["year"], last["year"] + 1):
            expected.append({"key": year, "label": str(year)})

    elif time_type == "Monthly":
        start_index = first["year"] * 12 + first["month"]
        end_index = last["year"] * 12 + last["month"]

        for index in range(start_index, end_index + 1):
            year = (index - 1) // 12
            month = ((index - 1) % 12) + 1
            expected.append(
                {"key": index, "label": f"{month:02d} - {year}"}
            )

    else:
        start_index = first["year"] * 4 + first["quarter"]
        end_index = last["year"] * 4 + last["quarter"]

        for index in range(start_index, end_index + 1):
            year = (index - 1) // 4
            quarter = ((index - 1) % 4) + 1
            expected.append(
                {"key": index, "label": f"{year} Q{quarter}"}
            )

    return expected


def repeated_runs_for_row(row, period_columns, threshold):
    values = []
    for column in period_columns:
        numeric = pd.to_numeric(
            pd.Series([row[column]]),
            errors="coerce"
        ).iloc[0]
        values.append(numeric)

    results = []
    i = 0

    while i < len(values):
        if pd.isna(values[i]):
            i += 1
            continue

        j = i + 1
        while (
            j < len(values)
            and not pd.isna(values[j])
            and values[j] == values[i]
        ):
            j += 1

        run_length = j - i

        if run_length >= threshold:
            results.append({
                "start_index": i,
                "end_index": j - 1,
                "length": run_length,
                "value": values[i]
            })

        i = j

    return results


def compress_missing_ranges(missing_indexes, period_columns):
    if not missing_indexes:
        return []

    ranges = []
    start = previous = missing_indexes[0]

    for current in missing_indexes[1:]:
        if current == previous + 1:
            previous = current
            continue

        ranges.append((start, previous))
        start = previous = current

    ranges.append((start, previous))

    output = []

    for start_index, end_index in ranges:
        start_label = str(period_columns[start_index])
        end_label = str(period_columns[end_index])

        output.append({
            "start_index": start_index,
            "end_index": end_index,
            "label": (
                start_label
                if start_index == end_index
                else f"{start_label} → {end_label}"
            ),
            "length": end_index - start_index + 1
        })

    return output


def get_flag_names(folder):
    names = []
    if not os.path.isdir(folder):
        return names

    for filename in os.listdir(folder):
        if filename.lower().endswith(IMAGE_EXTENSIONS):
            names.append(os.path.splitext(filename)[0])

    return sorted(set(names))


def find_flag_exact(entity, folder):
    if not os.path.isdir(folder):
        return None

    target = normalize_name(entity)

    for filename in os.listdir(folder):
        if not filename.lower().endswith(IMAGE_EXTENSIONS):
            continue

        stem = os.path.splitext(filename)[0]
        if normalize_name(stem) == target:
            return stem

    return None


def suggest_flag_match(entity, available_names):
    normalized_to_original = {
        normalize_name(name): name
        for name in available_names
    }

    close = get_close_matches(
        normalize_name(entity),
        list(normalized_to_original.keys()),
        n=1,
        cutoff=0.72
    )

    if close:
        return normalized_to_original[close[0]]

    return ""


def detect_correction_decimals(dataframe, period_columns):
    values = []

    for column in period_columns:
        numeric = pd.to_numeric(
            dataframe[column],
            errors="coerce"
        ).dropna()

        if not numeric.empty:
            values.extend(numeric.astype(float).tolist())

    if not values:
        return 2

    sample = values[:5000]

    for decimals in range(0, 7):
        matching = 0

        for value in sample:
            if abs(value - round(value, decimals)) < 10 ** (-(decimals + 4)):
                matching += 1

        if matching / len(sample) >= 0.95:
            return decimals

    return 6


def surrounding_values(df, row_index, start_index, end_index, period_columns):
    previous_value = None
    next_value = None

    for i in range(start_index - 1, -1, -1):
        value = pd.to_numeric(
            pd.Series([df.loc[row_index, period_columns[i]]]),
            errors="coerce"
        ).iloc[0]

        if not pd.isna(value):
            previous_value = float(value)
            break

    for i in range(end_index + 1, len(period_columns)):
        value = pd.to_numeric(
            pd.Series([df.loc[row_index, period_columns[i]]]),
            errors="coerce"
        ).iloc[0]

        if not pd.isna(value):
            next_value = float(value)
            break

    return previous_value, next_value


def linear_fill_values(count, previous_value, next_value, decimals):
    if (
        count <= 0
        or previous_value is None
        or next_value is None
    ):
        return None

    step = (next_value - previous_value) / (count + 1)

    return [
        round(previous_value + step * (i + 1), decimals)
        for i in range(count)
    ]


def make_corrected_workbook(
    uploaded_bytes,
    original_file_name,
    selected_sheet_name,
    corrected_df
):
    output = io.BytesIO()

    if original_file_name.lower().endswith(".csv"):
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            corrected_df.to_excel(
                writer,
                sheet_name="Corrected_Data",
                index=False
            )
    else:
        workbook = load_workbook(io.BytesIO(uploaded_bytes))
        worksheet = workbook[selected_sheet_name]

        for col_index, column_name in enumerate(
            corrected_df.columns,
            start=1
        ):
            worksheet.cell(
                row=1,
                column=col_index
            ).value = str(column_name)

        # Remove all existing data rows first so deleted dataframe rows
        # do not remain as stale rows in the workbook.
        if worksheet.max_row >= 2:
            worksheet.delete_rows(
                2,
                worksheet.max_row - 1
            )

        # Rewrite the corrected dataframe from row 2 onward.
        for row_offset, (_, data_row) in enumerate(
            corrected_df.iterrows(),
            start=2
        ):
            for col_index, column_name in enumerate(
                corrected_df.columns,
                start=1
            ):
                value = data_row[column_name]

                if pd.isna(value):
                    excel_value = None
                elif hasattr(value, "item"):
                    excel_value = value.item()
                else:
                    excel_value = value

                worksheet.cell(
                    row=row_offset,
                    column=col_index
                ).value = excel_value

        workbook.save(output)

    output.seek(0)
    return output.getvalue()


# =========================================================
# STEP 1 — UPLOAD
# =========================================================

st.header("1. Upload Excel")

uploaded_file = st.file_uploader(
    "Choose Excel or CSV file",
    type=["xlsx", "xls", "csv"]
)

if uploaded_file is None:
    st.info("Upload a file to begin.")
    st.stop()

uploaded_bytes = uploaded_file.getvalue()
file_name = uploaded_file.name

if file_name.lower().endswith(".csv"):
    df_source = pd.read_csv(
        io.BytesIO(uploaded_bytes),
        keep_default_na=False
    )
    sheet_name = "CSV"
else:
    excel_book = pd.ExcelFile(io.BytesIO(uploaded_bytes))
    if len(excel_book.sheet_names) == 1:
        sheet_name = excel_book.sheet_names[0]
        st.caption(f"Sheet: {sheet_name}")
    else:
        sheet_name = st.selectbox(
            "Sheet",
            excel_book.sheet_names
        )

    df_source = pd.read_excel(
        io.BytesIO(uploaded_bytes),
        sheet_name=sheet_name,
        keep_default_na=False
    )

detected_entity = detect_entity_column(df_source)
periods, time_type = detect_period_columns(df_source)

if not periods:
    st.error("No annual, monthly, or quarterly columns were detected.")
    st.stop()

entity_column = detected_entity

period_columns = [
    item["original_column"]
    for item in periods
]

file_key = hashlib.md5(
    (
        file_name
        + "|"
        + str(sheet_name)
        + "|"
        + str(len(uploaded_bytes))
    ).encode("utf-8")
).hexdigest()

if st.session_state.get("_simple_checker_file_key") != file_key:
    st.session_state["_simple_checker_file_key"] = file_key
    st.session_state["_simple_checker_df"] = df_source.copy()
    st.session_state["_simple_checker_checked"] = False
    st.session_state["_simple_checker_log"] = []

working_df = st.session_state["_simple_checker_df"].copy()

for column in period_columns:
    working_df[column] = working_df[column].astype("object")

st.session_state["_simple_checker_df"] = working_df

c1, c2, c3, c4 = st.columns(4)

c1.metric("Entities", len(working_df))
c2.metric("Periods", len(period_columns))
c3.metric("Time Type", time_type)
c4.metric(
    "Range",
    f"{period_columns[0]} → {period_columns[-1]}"
)

st.caption(
    f"Entity column detected automatically: {entity_column}"
)


# =========================================================
# ADVANCED SETTINGS
# =========================================================

with st.expander("Advanced Settings", expanded=False):

    a1, a2 = st.columns(2)

    with a1:
        data_type = st.selectbox(
            "Data Type",
            [
                "Regular Data",
                "Cumulative Totals"
            ],
            help=(
                "Use Cumulative Totals for data such as total COVID deaths. "
                "Repeated totals are allowed, but decreases are flagged."
            )
        )

        repeated_threshold = st.number_input(
            "Repeated Value Threshold",
            min_value=2,
            max_value=20,
            value=2,
            step=1
        )

        important_rank = st.number_input(
            "Important Rank",
            min_value=1,
            max_value=max(1, len(working_df)),
            value=min(20, len(working_df)),
            step=1
        )

    with a2:
        suspicious_jump_threshold = st.number_input(
            "Suspicious Jump %",
            min_value=100,
            max_value=10000,
            value=500,
            step=50
        )

        check_negative_values = st.checkbox(
            "Check Negative Values",
            value=True
        )

        flag_library_mode = st.selectbox(
            "Flag Check",
            [
                "Auto",
                "Countries / Historical",
                "US States",
                "Do Not Check Flags"
            ]
        )


# =========================================================
# STEP 2 — CHECK
# =========================================================

st.header("2. Check File")

check_clicked = st.button(
    "CHECK EXCEL",
    type="primary",
    use_container_width=True
)

if check_clicked:
    st.session_state["_simple_checker_checked"] = True

if not st.session_state.get("_simple_checker_checked", False):
    st.stop()

df = st.session_state["_simple_checker_df"].copy()
correction_decimals = detect_correction_decimals(
    df_source,
    period_columns
)

audit_rows = []

def add_audit(
    category,
    severity,
    entity="",
    period="",
    value="",
    details="",
    suggestion="",
    fixable=False
):
    audit_rows.append({
        "Category": category,
        "Severity": severity,
        "Entity": entity,
        "Period": period,
        "Value": value,
        "Details": details,
        "Suggestion": suggestion,
        "Fixable": fixable
    })

progress = st.progress(0)
status = st.empty()

# ---------------------------------------------------------
# Structure
# ---------------------------------------------------------

status.write("Checking structure...")
progress.progress(10)

blank_entity_mask = (
    df[entity_column].isna()
    | df[entity_column].astype(str).str.strip().eq("")
)

for index in df.index[blank_entity_mask]:
    add_audit(
        "Blank Entity Name",
        "Error",
        details=f"Blank entity name on Excel row {index + 2}."
    )

entity_clean = df[entity_column].astype(str).str.strip()

duplicate_mask = (
    entity_clean.ne("")
    & entity_clean.duplicated(keep=False)
)

for entity_name, group in df[duplicate_mask].groupby(
    entity_clean[duplicate_mask]
):
    excel_rows = [int(index) + 2 for index in group.index]
    add_audit(
        "Duplicate Entity",
        "Error",
        entity=entity_name,
        details=f"Appears on Excel rows {excel_rows}."
    )

full_duplicate_mask = df.duplicated(keep=False)

for index in df.index[full_duplicate_mask]:
    add_audit(
        "Duplicate Row",
        "Error",
        entity=str(df.loc[index, entity_column]),
        details=f"Entire row duplicated. Excel row {index + 2}."
    )

# ---------------------------------------------------------
# Date / period
# ---------------------------------------------------------

status.write("Checking dates...")
progress.progress(22)

period_key_map = {}

for period in periods:
    period_key_map.setdefault(
        period["sort_key"],
        []
    ).append(
        str(period["original_column"])
    )

for labels in period_key_map.values():
    if len(labels) > 1:
        add_audit(
            "Duplicate Period",
            "Error",
            period=" / ".join(labels),
            details="The same period appears more than once."
        )

actual_keys = [item["sort_key"] for item in periods]

if actual_keys != sorted(actual_keys):
    add_audit(
        "Periods Out of Order",
        "Warning",
        details="Time columns are not arranged chronologically."
    )

expected = expected_period_keys(periods, time_type)
existing_keys = set(actual_keys)

for item in expected:
    if item["key"] not in existing_keys:
        add_audit(
            "Missing Period",
            "Warning",
            period=item["label"],
            details=f"{item['label']} is missing from the time-series columns."
        )

# ---------------------------------------------------------
# Missing / invalid
# ---------------------------------------------------------

status.write("Checking missing and invalid data...")
progress.progress(36)

for row_index, row in df.iterrows():
    entity = str(row[entity_column]).strip()
    converted = []

    for column in period_columns:
        raw_value = row[column]

        if is_true_blank(raw_value):
            converted.append(("blank", None, raw_value))
            continue

        numeric_value = pd.to_numeric(
            pd.Series([raw_value]),
            errors="coerce"
        ).iloc[0]

        if pd.isna(numeric_value):
            converted.append(("invalid", None, raw_value))
        else:
            converted.append(("number", float(numeric_value), raw_value))

    valid_positions = [
        i for i, item in enumerate(converted)
        if item[0] == "number"
    ]

    for i, item in enumerate(converted):
        kind, numeric_value, raw_value = item

        if kind == "invalid":
            add_audit(
                "Non-Numeric Value",
                "Error",
                entity=entity,
                period=str(period_columns[i]),
                value=str(raw_value),
                details=f"Expected a number on Excel row {row_index + 2}.",
                suggestion="Replace with the correct number."
            )

    if valid_positions:
        first_valid = valid_positions[0]
        last_valid = valid_positions[-1]

        missing_inside = [
            i for i in range(first_valid + 1, last_valid)
            if converted[i][0] == "blank"
        ]

        for missing_range in compress_missing_ranges(
            missing_inside,
            period_columns
        ):
            add_audit(
                "Internal Gap",
                "Warning",
                entity=entity,
                period=missing_range["label"],
                details=(
                    f"{missing_range['length']} missing period(s) "
                    f"inside the series. Excel row {row_index + 2}."
                ),
                suggestion="Series Fill is available when there is a real value before and after.",
                fixable=True
            )

# ---------------------------------------------------------
# Number checks
# ---------------------------------------------------------

status.write("Checking numbers...")
progress.progress(52)

for row_index, row in df.iterrows():
    entity = str(row[entity_column]).strip()

    runs = repeated_runs_for_row(
        row,
        period_columns,
        int(repeated_threshold)
    )

    for run in runs:
        start_column = str(period_columns[run["start_index"]])
        end_column = str(period_columns[run["end_index"]])
        value = run["value"]

        if float(value) == 0:
            add_audit(
                "Zero Run",
                "Review",
                entity=entity,
                period=f"{start_column} → {end_column}",
                value=value,
                details=f"Zero appears in {run['length']} consecutive periods.",
                suggestion="Review whether zeros are real values or placeholders."
            )

        elif data_type == "Regular Data":
            add_audit(
                "Repeated Consecutive Value",
                "Review",
                entity=entity,
                period=f"{start_column} → {end_column}",
                value=value,
                details=f"Same non-zero value appears in {run['length']} consecutive periods.",
                suggestion="Review whether this is genuine or accidental forward-fill."
            )

    previous_numeric = None
    previous_column = None

    for column in period_columns:
        current_numeric = pd.to_numeric(
            pd.Series([row[column]]),
            errors="coerce"
        ).iloc[0]

        if pd.isna(current_numeric):
            continue

        current_numeric = float(current_numeric)

        if check_negative_values and current_numeric < 0:
            add_audit(
                "Negative Value",
                "Review",
                entity=entity,
                period=str(column),
                value=current_numeric,
                details=f"Negative value on Excel row {row_index + 2}.",
                suggestion="Review whether negatives are valid for this dataset."
            )

        if (
            previous_numeric is not None
            and previous_numeric != 0
        ):
            change_percent = abs(
                (
                    current_numeric - previous_numeric
                )
                / abs(previous_numeric)
                * 100
            )

            if change_percent > float(suspicious_jump_threshold):
                add_audit(
                    "Suspicious Jump",
                    "Review",
                    entity=entity,
                    period=f"{previous_column} → {column}",
                    value=current_numeric,
                    details=(
                        f"Changed by {change_percent:,.1f}% "
                        f"from {previous_numeric:,.6g} "
                        f"to {current_numeric:,.6g}."
                    ),
                    suggestion="Review this large change."
                )

        if (
            data_type == "Cumulative Totals"
            and previous_numeric is not None
            and current_numeric < previous_numeric
        ):
            add_audit(
                "Cumulative Value Decreased",
                "Error",
                entity=entity,
                period=f"{previous_column} → {column}",
                value=current_numeric,
                details=(
                    f"Cumulative value fell from "
                    f"{previous_numeric:,.6g} "
                    f"to {current_numeric:,.6g}."
                ),
                suggestion="Cumulative totals normally should not decrease."
            )

        previous_numeric = current_numeric
        previous_column = str(column)

# ---------------------------------------------------------
# Priority ranks
# ---------------------------------------------------------

status.write("Checking important entities...")
progress.progress(66)

best_rank_by_entity = {}

for period in periods:
    column = period["original_column"]
    numeric_values = pd.to_numeric(
        df[column],
        errors="coerce"
    )
    ranks = numeric_values.rank(
        method="min",
        ascending=False
    )

    for row_index, rank_value in ranks.items():
        if pd.isna(rank_value):
            continue

        entity_name = str(
            df.loc[row_index, entity_column]
        ).strip()

        if not entity_name:
            continue

        rank_number = int(rank_value)
        current_best = best_rank_by_entity.get(entity_name)

        if (
            current_best is None
            or rank_number < current_best
        ):
            best_rank_by_entity[entity_name] = rank_number

# Important start/end coverage
for row_index, row in df.iterrows():
    entity_name = str(row[entity_column]).strip()

    if not entity_name:
        continue

    best_rank = best_rank_by_entity.get(entity_name)

    if (
        best_rank is None
        or best_rank > int(important_rank)
    ):
        continue

    numeric_series = pd.to_numeric(
        row[period_columns],
        errors="coerce"
    )

    valid_positions = [
        i for i, value in enumerate(numeric_series.tolist())
        if not pd.isna(value)
    ]

    if not valid_positions:
        continue

    leading_missing = list(range(0, valid_positions[0]))
    trailing_missing = list(
        range(valid_positions[-1] + 1, len(period_columns))
    )

    for missing_range in (
        compress_missing_ranges(leading_missing, period_columns)
        + compress_missing_ranges(trailing_missing, period_columns)
    ):
        add_audit(
            "Important Coverage Gap",
            "Warning",
            entity=entity_name,
            period=missing_range["label"],
            details=(
                f"{missing_range['length']} missing period(s) "
                f"for an entity that reached Top {int(important_rank)}."
            ),
            suggestion="Review whether reliable data can be added."
        )

# ---------------------------------------------------------
# Flags
# ---------------------------------------------------------

status.write("Checking flags...")
progress.progress(78)

flag_library_available = (
    os.path.isdir(COUNTRY_FLAG_LIBRARY)
    or os.path.isdir(STATE_FLAG_LIBRARY)
)

if flag_library_mode == "Do Not Check Flags":
    flag_check_status = "Flag checking turned off."

elif not flag_library_available:
    flag_check_status = (
        "Flag library is not connected on this online app, "
        "so flag checking was skipped."
    )

else:
    country_names = get_flag_names(COUNTRY_FLAG_LIBRARY)
    state_names = get_flag_names(STATE_FLAG_LIBRARY)

    entities = (
        df[entity_column]
        .dropna()
        .astype(str)
        .str.strip()
    )
    entities = entities[entities.ne("")]

    if flag_library_mode == "Countries / Historical":
        selected_folder = COUNTRY_FLAG_LIBRARY
        available_names = country_names
        selected_label = "Countries / Historical"

    elif flag_library_mode == "US States":
        selected_folder = STATE_FLAG_LIBRARY
        available_names = state_names
        selected_label = "US States"

    else:
        country_match_count = sum(
            find_flag_exact(entity, COUNTRY_FLAG_LIBRARY) is not None
            for entity in entities
        )
        state_match_count = sum(
            find_flag_exact(entity, STATE_FLAG_LIBRARY) is not None
            for entity in entities
        )

        if state_match_count > country_match_count:
            selected_folder = STATE_FLAG_LIBRARY
            available_names = state_names
            selected_label = "US States"
        else:
            selected_folder = COUNTRY_FLAG_LIBRARY
            available_names = country_names
            selected_label = "Countries / Historical"

    flag_check_status = f"Flag library used: {selected_label}"

    for entity in sorted(set(entities)):
        if find_flag_exact(entity, selected_folder) is not None:
            continue

        suggestion = suggest_flag_match(entity, available_names)

        if suggestion:
            add_audit(
                "Flag Name Mismatch",
                "Review",
                entity=entity,
                details=f"No exact flag filename matches '{entity}'.",
                suggestion=f"Possible match: {suggestion}"
            )
        else:
            add_audit(
                "Missing Flag",
                "Warning",
                entity=entity,
                details=f"No matching flag found in {selected_label} library."
            )

# ---------------------------------------------------------
# Completeness
# ---------------------------------------------------------

status.write("Calculating completeness...")
progress.progress(90)

completeness_rows = []

for _, row in df.iterrows():
    entity = str(row[entity_column]).strip()

    numeric_series = pd.to_numeric(
        row[period_columns],
        errors="coerce"
    )

    valid_count = int(numeric_series.notna().sum())
    missing_count = len(period_columns) - valid_count

    valid_positions = [
        i for i, value in enumerate(numeric_series.tolist())
        if not pd.isna(value)
    ]

    first_available = (
        str(period_columns[valid_positions[0]])
        if valid_positions
        else ""
    )

    last_available = (
        str(period_columns[valid_positions[-1]])
        if valid_positions
        else ""
    )

    completeness_rows.append({
        "Entity": entity,
        "Completeness %": round(
            valid_count / len(period_columns) * 100,
            1
        ) if period_columns else 0,
        "Observations": valid_count,
        "Missing": missing_count,
        "First Available": first_available,
        "Last Available": last_available
    })

# Priority fields
for finding in audit_rows:
    entity_name = str(finding.get("Entity", "")).strip()
    best_rank = best_rank_by_entity.get(entity_name)

    finding["Best Rank"] = (
        best_rank if best_rank is not None else ""
    )

    if (
        best_rank is not None
        and best_rank <= int(important_rank)
    ):
        finding["Priority"] = "HIGH PRIORITY"
    elif entity_name:
        finding["Priority"] = "Lower Priority"
    else:
        finding["Priority"] = "General"

audit_df = pd.DataFrame(audit_rows)
if not audit_df.empty:
    audit_df = audit_df.drop_duplicates(
        subset=[
            "Priority", "Severity", "Category", "Entity",
            "Best Rank", "Period", "Value", "Details", "Suggestion"
        ],
        keep="first"
    ).reset_index(drop=True)
completeness_df = pd.DataFrame(completeness_rows)

if not audit_df.empty:
    order = {
        "HIGH PRIORITY": 0,
        "General": 1,
        "Lower Priority": 2
    }

    audit_df["_PriorityOrder"] = (
        audit_df["Priority"]
        .map(order)
        .fillna(9)
    )

    audit_df = (
        audit_df
        .sort_values(
            ["_PriorityOrder", "Best Rank", "Severity", "Category", "Entity"],
            na_position="last"
        )
        .drop(columns=["_PriorityOrder"])
        .reset_index(drop=True)
    )

progress.progress(100)
status.write("Check complete.")


# =========================================================
# STEP 3 — SIMPLE RESULT
# =========================================================

st.header("3. Problems")

if audit_df.empty:
    st.success("READY FOR BAR CHART RACE — No problems found.")
else:
    flag_categories = ["Missing Flag", "Flag Name Mismatch"]

    must_fix = int(
        (
            audit_df["Severity"].eq("Error")
            & ~audit_df["Category"].isin(flag_categories)
        ).sum()
    )

    should_review = int(
        (
            audit_df["Severity"].isin(["Warning", "Review"])
            & ~audit_df["Category"].isin(flag_categories)
        ).sum()
    )

    flag_problems = int(
        audit_df["Category"].isin(flag_categories).sum()
    )

    missing_problems = int(
        audit_df["Category"].isin(
            ["Internal Gap", "Important Coverage Gap", "Missing Period"]
        ).sum()
    )

    s1, s2, s3, s4 = st.columns(4)

    s1.metric("Must Fix", must_fix)
    s2.metric("Needs Review", should_review)
    s3.metric("Missing Data", missing_problems)
    s4.metric("Flag Problems", flag_problems)

    if must_fix > 0:
        st.error(
            f"{must_fix} problem(s) should be fixed before making the video."
        )
    elif should_review > 0:
        st.warning(
            "No critical data errors, but some items should be reviewed."
        )
    else:
        st.success("Data checks look good.")

# =========================================================
# STEP 4 — FIX
# =========================================================

st.header("4. Review All Problems")

# Entity-first review state
if "_country_review_decisions" not in st.session_state:
    st.session_state["_country_review_decisions"] = {}
if "_completed_countries" not in st.session_state:
    st.session_state["_completed_countries"] = []

def _problem_id(entity_name, finding):
    return "||".join([
        str(entity_name),
        str(finding.get("Category", "")),
        str(finding.get("Period", "")),
        str(finding.get("Details", ""))
    ])

def _entity_row(entity_name):
    matches = df.index[
        df[entity_column].astype(str).str.strip().eq(str(entity_name).strip())
    ].tolist()
    return matches[0] if matches else None

def _context_rows(entity_name, period_text):
    """Show one value before and after the affected numerical period/range."""
    row_index = _entity_row(entity_name)
    if row_index is None:
        return []

    labels = [str(c) for c in period_columns]
    pieces = [p.strip() for p in str(period_text).split("→")]
    indexes = [labels.index(p) for p in pieces if p in labels]
    if not indexes:
        return []

    start = max(0, min(indexes) - 1)
    end = min(len(labels) - 1, max(indexes) + 1)
    rows = []

    for i in range(start, end + 1):
        col = period_columns[i]
        raw = st.session_state["_simple_checker_df"].loc[row_index, col]
        numeric = pd.to_numeric(pd.Series([raw]), errors="coerce").iloc[0]
        rows.append({
            "Period": str(col),
            "Value": "Missing" if pd.isna(numeric) else numeric
        })
    return rows

def _series_fill_plan(entity_name, period_text):
    row_index = _entity_row(entity_name)
    if row_index is None:
        return None

    labels = [str(c) for c in period_columns]
    if "→" in str(period_text):
        start_label, end_label = [p.strip() for p in str(period_text).split("→", 1)]
    else:
        start_label = end_label = str(period_text).strip()

    if start_label not in labels or end_label not in labels:
        return None

    start_index = labels.index(start_label)
    end_index = labels.index(end_label)
    prev_value, next_value = surrounding_values(
        st.session_state["_simple_checker_df"],
        row_index,
        start_index,
        end_index,
        period_columns
    )
    if prev_value is None or next_value is None:
        return None

    cols = period_columns[start_index:end_index + 1]
    values = linear_fill_values(
        len(cols), prev_value, next_value, int(correction_decimals)
    )
    if values is None:
        return None

    return {
        "row_index": row_index,
        "columns": cols,
        "values": values,
        "before": prev_value,
        "after": next_value
    }

if not audit_df.empty:
    country_audit = audit_df[
        audit_df["Entity"].astype(str).str.strip().ne("")
    ].copy()

    blank_audit = audit_df[
        audit_df["Category"].astype(str).eq("Blank Entity Name")
    ].copy()

    blank_review_items = []
    for _, blank_finding in blank_audit.reset_index(drop=True).iterrows():
        details_text = str(blank_finding.get("Details", ""))
        import re
        row_match = re.search(r"Excel row\s+(\d+)", details_text)
        excel_row = row_match.group(1) if row_match else "?"
        blank_review_items.append({
            "label": f"BLANK ENTITY — Excel Row {excel_row}",
            "excel_row": int(excel_row) if str(excel_row).isdigit() else None,
            "finding": blank_finding
        })

    # Priority = any entity that entered selected Top N at least once.
    priority_countries = []
    other_countries = []
    for name in country_audit["Entity"].astype(str).drop_duplicates():
        rank = best_rank_by_entity.get(name)
        if rank is not None and rank <= int(important_rank):
            priority_countries.append(name)
        else:
            other_countries.append(name)

    priority_countries.sort(key=lambda x: best_rank_by_entity.get(x, 999999))
    other_countries.sort()

    completed_set = set(st.session_state["_completed_countries"])

    priority_remaining = [
        c for c in priority_countries
        if c not in completed_set
    ]
    priority_completed = [
        c for c in priority_countries
        if c in completed_set
    ]

    m1, m2, m3 = st.columns(3)
    m1.metric(f"Entities Ever in Top {int(important_rank)}", len(priority_countries))
    m2.metric("Priority Remaining", len(priority_remaining))
    m3.metric("Priority Completed", len(priority_completed))

    if blank_review_items:
        st.error(
            f"{len(blank_review_items)} blank-entity row(s) need review. "
            "They are listed in this main review section so they cannot be missed."
        )

    if priority_countries:
        st.info(
            f"Top {int(important_rank)} entities with detected problems: "
            + ", ".join(priority_countries)
        )


    def _suspicious_jump_plan(entity_name, period_text):
        row_index = _entity_row(entity_name)
        if row_index is None or "→" not in str(period_text):
            return None
        labels = [str(c) for c in period_columns]
        left_label, bad_label = [p.strip() for p in str(period_text).split("→", 1)]
        if left_label not in labels or bad_label not in labels:
            return None
        left_index = labels.index(left_label)
        bad_index = labels.index(bad_label)
        if bad_index != left_index + 1 or bad_index + 1 >= len(period_columns):
            return None
        work = st.session_state["_simple_checker_df"]
        left = pd.to_numeric(pd.Series([work.loc[row_index, period_columns[left_index]]]), errors="coerce").iloc[0]
        old = pd.to_numeric(pd.Series([work.loc[row_index, period_columns[bad_index]]]), errors="coerce").iloc[0]
        right = pd.to_numeric(pd.Series([work.loc[row_index, period_columns[bad_index + 1]]]), errors="coerce").iloc[0]
        if pd.isna(left) or pd.isna(old) or pd.isna(right):
            return None
        proposed = round((float(left) + float(right)) / 2, int(correction_decimals))
        return {
            "row_index": row_index, "column": period_columns[bad_index],
            "old": float(old), "proposed": proposed,
            "before_period": str(period_columns[left_index]), "before": float(left),
            "after_period": str(period_columns[bad_index + 1]), "after": float(right)
        }


    def _range_interpolation_plan(entity_name, period_text):
        row_index = _entity_row(entity_name)
        if row_index is None:
            return None

        labels = [str(c) for c in period_columns]
        parts = [p.strip() for p in str(period_text).split("→")]
        if not parts or any(p not in labels for p in parts):
            return None

        start_index = labels.index(parts[0])
        end_index = labels.index(parts[-1])
        if end_index < start_index:
            start_index, end_index = end_index, start_index

        work = st.session_state["_simple_checker_df"]
        before, after = surrounding_values(
            work, row_index, start_index, end_index, period_columns
        )
        if before is None or after is None:
            return None

        values = linear_fill_values(
            end_index - start_index + 1,
            before,
            after,
            int(correction_decimals)
        )
        if values is None:
            return None

        return {
            "row_index": row_index,
            "columns": period_columns[start_index:end_index + 1],
            "values": values,
            "before": before,
            "after": after,
            "start_index": start_index,
            "end_index": end_index
        }


    def _affected_period_columns(period_text):
        labels = [str(c) for c in period_columns]
        text = str(period_text).strip()
        if not text:
            return []
        parts = [p.strip() for p in text.split("→") if p.strip()]
        if not parts or any(p not in labels for p in parts):
            return []
        start_i = labels.index(parts[0])
        end_i = labels.index(parts[-1])
        if end_i < start_i:
            start_i, end_i = end_i, start_i
        return list(period_columns[start_i:end_i + 1])

    def _entity_completion_label(entity_name, findings):
        decisions = []
        for _, finding in findings.iterrows():
            pid = _problem_id(entity_name, finding)
            decisions.append(
                st.session_state["_country_review_decisions"].get(pid, "Pending")
            )
        if decisions and all(d == "Ignored" for d in decisions):
            return "Reviewed — Kept Original"
        if "Ignored" in decisions and "Fixed" in decisions:
            return "Reviewed — Fixed + Kept Original"
        if decisions and all(d == "Fixed" for d in decisions):
            return "Fixed"
        return "Reviewed"

    def _entity_review_status(entity_findings):
        categories = set(entity_findings["Category"].astype(str).tolist())
        severities = set(entity_findings["Severity"].astype(str).tolist())

        must_fix_categories = {
            "Non-Numeric Value", "Negative Value", "Duplicate Entity",
            "Duplicate Row", "Blank Entity Name", "Duplicate Period",
            "Periods Out of Order", "Cumulative Value Decreased",
            "Important Coverage Gap"
        }

        # Hard errors remain red, even if an interpolation can be suggested.
        if "Error" in severities or categories.intersection(must_fix_categories):
            return "🔴 MUST FIX"

        for _, r in entity_findings.iterrows():
            category = str(r.get("Category", ""))
            period_text = str(r.get("Period", "")).strip()
            if category == "Internal Gap":
                if _series_fill_plan(str(r.get("Entity", "")), period_text) is not None:
                    return "🟡 EASY FIX"
            elif category == "Suspicious Jump":
                if _suspicious_jump_plan(str(r.get("Entity", "")), period_text) is not None:
                    return "🟡 EASY FIX"
            elif category in {"Repeated Consecutive Value", "Zero Run"}:
                if _range_interpolation_plan(str(r.get("Entity", "")), period_text) is not None:
                    return "🟡 EASY FIX"

        return "🟡 REVIEW"

    st.subheader("All Problems Needing Review")
    st.caption("Top-ranked problems are shown first. Every lower-priority and structural problem is also listed below.")

    for country_name in priority_remaining:
        findings = country_audit[
            country_audit["Entity"].astype(str).eq(country_name)
        ].reset_index(drop=True)

        best_rank = best_rank_by_entity.get(country_name)
        rank_text = f" • Best Rank #{best_rank}" if best_rank is not None else ""

        status_text = _entity_review_status(findings)
        with st.expander(
            f"{country_name} — {len(findings)} problem(s){rank_text}     {status_text}",
            expanded=False
        ):
            for problem_number, finding in findings.iterrows():
                category = str(finding["Category"])
                period_text = str(finding.get("Period", "")).strip()
                pid = _problem_id(country_name, finding)
                decision = st.session_state["_country_review_decisions"].get(
                    pid, "Pending"
                )

                st.markdown(f"### {problem_number + 1}. {category}")
                if period_text:
                    st.write(f"**Period:** {period_text}")

                # For numerical problems, show the affected values plus one before/after.
                context = _context_rows(country_name, period_text)
                if context:
                    st.dataframe(
                        pd.DataFrame(context),
                        use_container_width=True,
                        hide_index=True
                    )

                details = str(finding.get("Details", "")).strip()
                suggestion = str(finding.get("Suggestion", "")).strip()
                if details:
                    st.write(details)
                if suggestion:
                    st.caption(suggestion)

                plan = None
                jump_plan = None
                interpolation_plan = None

                if category == "Internal Gap":
                    plan = _series_fill_plan(country_name, period_text)
                elif category == "Suspicious Jump":
                    jump_plan = _suspicious_jump_plan(country_name, period_text)
                elif category in {
                    "Repeated Consecutive Value",
                    "Zero Run",
                    "Non-Numeric Value",
                    "Negative Value"
                }:
                    interpolation_plan = _range_interpolation_plan(
                        country_name, period_text
                    )

                # Root-error protection:
                # if this Suspicious Jump touches a period already marked as a stronger
                # root problem, do not suggest a jump fix based on that bad value.
                if category == "Suspicious Jump":
                    root_periods = {
                        str(r.get("Period", "")).strip()
                        for _, r in findings.iterrows()
                        if str(r.get("Category", "")) in {
                            "Negative Value",
                            "Non-Numeric Value",
                            "Internal Gap",
                            "Important Coverage Gap"
                        }
                    }
                    jump_parts = [p.strip() for p in period_text.split("→")]
                    if any(p in root_periods for p in jump_parts):
                        jump_plan = None

                if plan is not None and decision == "Pending":
                    st.write(
                        f"**Before:** {plan['before']:,.0f}   |   "
                        f"**After:** {plan['after']:,.0f}"
                    )
                    st.write("**Proposed Series Fill:**")
                    st.dataframe(
                        pd.DataFrame({
                            "Period": [str(c) for c in plan["columns"]],
                            "Proposed Value": plan["values"]
                        }),
                        use_container_width=True,
                        hide_index=True
                    )

                if interpolation_plan is not None and decision == "Pending":
                    st.warning(
                        "Suggested Series Fill is available because valid values "
                        "exist before and after this problem range."
                    )
                    st.write(
                        f"**Before:** {interpolation_plan['before']:,.0f}   |   "
                        f"**After:** {interpolation_plan['after']:,.0f}"
                    )
                    st.dataframe(
                        pd.DataFrame({
                            "Period": [str(c) for c in interpolation_plan["columns"]],
                            "Proposed Value": interpolation_plan["values"]
                        }),
                        use_container_width=True,
                        hide_index=True
                    )

                if (
                    interpolation_plan is None
                    and category == "Important Coverage Gap"
                    and decision == "Pending"
                ):
                    labels = [str(c) for c in period_columns]
                    parts = [p.strip() for p in period_text.split("→") if p.strip()]
                    if parts and all(p in labels for p in parts):
                        start_i = labels.index(parts[0])
                        end_i = labels.index(parts[-1])
                        row_i = _entity_row(country_name)

                        if start_i == 0 and end_i + 1 < len(period_columns):
                            next_col = period_columns[end_i + 1]
                            next_val = pd.to_numeric(
                                pd.Series([
                                    st.session_state["_simple_checker_df"].loc[
                                        row_i, next_col
                                    ]
                                ]),
                                errors="coerce"
                            ).iloc[0]
                            if not pd.isna(next_val):
                                st.info(
                                    f"First available value after the gap: "
                                    f"**{next_col} = {float(next_val):,.0f}**. "
                                    "Automatic fix is unavailable because there is "
                                    "no value before the missing range."
                                )

                        elif end_i == len(period_columns) - 1 and start_i > 0:
                            prev_col = period_columns[start_i - 1]
                            prev_val = pd.to_numeric(
                                pd.Series([
                                    st.session_state["_simple_checker_df"].loc[
                                        row_i, prev_col
                                    ]
                                ]),
                                errors="coerce"
                            ).iloc[0]
                            if not pd.isna(prev_val):
                                st.info(
                                    f"Last available value before the gap: "
                                    f"**{prev_col} = {float(prev_val):,.0f}**. "
                                    "Automatic fix is unavailable because there is "
                                    "no value after the missing range."
                                )

                if jump_plan is not None and decision == "Pending":
                    st.warning(
                        f"Easy fix suggestion: replace **{jump_plan['column']}** "
                        f"from **{jump_plan['old']:,.6g}** to **{jump_plan['proposed']:,.6g}**."
                    )
                    st.write(
                        f"{jump_plan['before_period']}: **{jump_plan['before']:,.6g}**  →  "
                        f"{jump_plan['column']}: **{jump_plan['proposed']:,.6g}**  →  "
                        f"{jump_plan['after_period']}: **{jump_plan['after']:,.6g}**"
                    )

                c1, c2, c3 = st.columns(3)

                if decision == "Pending":
                    if plan is not None:
                        if c1.button(
                            "Fix — Series Fill",
                            key=f"country_fix_{abs(hash(pid))}",
                            use_container_width=True
                        ):
                            corrected = st.session_state["_simple_checker_df"].copy()
                            for col, value in zip(plan["columns"], plan["values"]):
                                corrected.loc[plan["row_index"], col] = value
                            st.session_state["_simple_checker_df"] = corrected
                            st.session_state["_country_review_decisions"][pid] = "Fixed"
                            st.session_state["_simple_checker_log"].append({
                                "Entity": country_name,
                                "Problem": category,
                                "Period": period_text,
                                "Method": "Series Fill",
                                "Updated Cells": len(plan["columns"])
                            })
                            st.rerun()
                    elif interpolation_plan is not None:
                        if c1.button(
                            "Apply Suggested Series Fill",
                            key=f"interp_fix_{abs(hash(pid))}",
                            use_container_width=True
                        ):
                            corrected = st.session_state["_simple_checker_df"].copy()
                            old_values = []
                            for col, value in zip(
                                interpolation_plan["columns"],
                                interpolation_plan["values"]
                            ):
                                old_values.append(
                                    corrected.loc[
                                        interpolation_plan["row_index"], col
                                    ]
                                )
                                corrected.loc[
                                    interpolation_plan["row_index"], col
                                ] = value

                            st.session_state["_simple_checker_df"] = corrected
                            st.session_state["_country_review_decisions"][pid] = "Fixed"
                            st.session_state["_simple_checker_log"].append({
                                "Entity": country_name,
                                "Problem": category,
                                "Period": period_text,
                                "Method": "Suggested Series Fill",
                                "Old Value": str(old_values),
                                "New Value": str(interpolation_plan["values"])
                            })
                            st.rerun()

                    elif jump_plan is not None:
                        if c1.button(
                            "Apply Suggested Fix",
                            key=f"jump_fix_{abs(hash(pid))}",
                            use_container_width=True
                        ):
                            corrected = st.session_state["_simple_checker_df"].copy()
                            corrected.loc[
                                jump_plan["row_index"],
                                jump_plan["column"]
                            ] = jump_plan["proposed"]
                            st.session_state["_simple_checker_df"] = corrected
                            st.session_state["_country_review_decisions"][pid] = "Fixed"
                            st.session_state["_simple_checker_log"].append({
                                "Entity": country_name,
                                "Problem": category,
                                "Period": period_text,
                                "Method": "Interpolated Suspicious Jump",
                                "Old Value": jump_plan["old"],
                                "New Value": jump_plan["proposed"]
                            })
                            st.rerun()
                    else:
                        manual_cols = _affected_period_columns(period_text)
                        if manual_cols:
                            if c1.button(
                                "Manual Entry",
                                key=f"manual_open_{abs(hash(pid))}",
                                use_container_width=True
                            ):
                                st.session_state[f"_manual_open_{abs(hash(pid))}"] = True
                        else:
                            c1.button(
                                "Manual Fix Required",
                                key=f"manual_{abs(hash(pid))}",
                                disabled=True,
                                help="This structural problem cannot be corrected with a single numeric entry.",
                                use_container_width=True
                            )

                    if c2.button(
                        "Ignore / Keep Original",
                        key=f"country_ignore_{abs(hash(pid))}",
                        use_container_width=True
                    ):
                        st.session_state["_country_review_decisions"][pid] = "Ignored"
                        st.rerun()
                else:
                    c1.success(
                        "Reviewed — Kept Original"
                        if decision == "Ignored"
                        else "Fixed"
                    )
                    if c2.button(
                        "Undo",
                        key=f"country_undo_{abs(hash(pid))}",
                        use_container_width=True
                    ):
                        st.session_state["_country_review_decisions"][pid] = "Pending"
                        st.rerun()

                manual_key = f"_manual_open_{abs(hash(pid))}"
                if decision == "Pending" and st.session_state.get(manual_key, False):
                    manual_cols = _affected_period_columns(period_text)
                    if manual_cols:
                        st.write("**Manual Entry**")
                        st.caption(
                            "Enter the value(s) you verified from your source. "
                            "The app will put them into the corrected Excel file."
                        )
                        manual_values = {}
                        row_i = _entity_row(country_name)
                        for col in manual_cols:
                            current_raw = st.session_state["_simple_checker_df"].loc[
                                row_i, col
                            ]
                            current_num = pd.to_numeric(
                                pd.Series([current_raw]), errors="coerce"
                            ).iloc[0]
                            default_value = (
                                float(current_num) if not pd.isna(current_num) else 0.0
                            )
                            manual_values[col] = st.number_input(
                                f"{col}",
                                value=default_value,
                                key=f"manual_value_{abs(hash(pid))}_{str(col)}"
                            )

                        mc1, mc2 = st.columns(2)
                        if mc1.button(
                            "Apply Manual Entry",
                            key=f"manual_apply_{abs(hash(pid))}",
                            type="primary",
                            use_container_width=True
                        ):
                            corrected = st.session_state["_simple_checker_df"].copy()
                            old_values = {}
                            for col, value in manual_values.items():
                                old_values[str(col)] = corrected.loc[row_i, col]
                                corrected.loc[row_i, col] = value
                            st.session_state["_simple_checker_df"] = corrected
                            st.session_state["_country_review_decisions"][pid] = "Fixed"
                            st.session_state["_simple_checker_log"].append({
                                "Entity": country_name,
                                "Problem": category,
                                "Period": period_text,
                                "Method": "Manual Entry",
                                "Old Value": str(old_values),
                                "New Value": str({
                                    str(k): v for k, v in manual_values.items()
                                })
                            })
                            st.session_state[manual_key] = False
                            st.rerun()

                        if mc2.button(
                            "Cancel Manual Entry",
                            key=f"manual_cancel_{abs(hash(pid))}",
                            use_container_width=True
                        ):
                            st.session_state[manual_key] = False
                            st.rerun()

                st.divider()

            pending = 0
            for _, finding in findings.iterrows():
                pid = _problem_id(country_name, finding)
                if st.session_state["_country_review_decisions"].get(
                    pid, "Pending"
                ) == "Pending":
                    pending += 1

            if pending == 0:
                if st.button(
                    f"ENTITY COMPLETE — {country_name}",
                    key=f"complete_{normalize_name(country_name)}",
                    type="primary",
                    use_container_width=True
                ):
                    if country_name not in st.session_state["_completed_countries"]:
                        st.session_state["_completed_countries"].append(country_name)
                    st.rerun()
            else:
                st.caption(
                    f"{pending} problem(s) remaining. Fix or Ignore each one to complete this entity."
                )

    if priority_completed:
        with st.expander(
            f"Completed Priority Entities ({len(priority_completed)})",
            expanded=False
        ):
            for country_name in priority_completed:
                completed_findings = country_audit[
                    country_audit["Entity"].astype(str).eq(country_name)
                ].reset_index(drop=True)
                completion_label = _entity_completion_label(
                    country_name, completed_findings
                )
                cc1, cc2 = st.columns([4, 1])
                cc1.write(f"✅ {country_name} — **{completion_label}**")
                if cc2.button(
                    "Reopen",
                    key=f"reopen_{normalize_name(country_name)}"
                ):
                    st.session_state["_completed_countries"].remove(country_name)
                    st.rerun()

    if blank_review_items:
        st.markdown("### Structural / Unnamed Rows Needing Review")

        for blank_item in blank_review_items:
            blank_finding = blank_item["finding"]
            blank_label = blank_item["label"]
            blank_row = blank_item["excel_row"]
            blank_pid = (
                "BLANK||" + str(blank_row) + "||"
                + str(blank_finding.get("Details", ""))
            )
            blank_decision = st.session_state[
                "_country_review_decisions"
            ].get(blank_pid, "Pending")

            if blank_decision == "Pending":
                with st.expander(
                    f"{blank_label}     🔴 MUST FIX",
                    expanded=False
                ):
                    st.write(str(blank_finding.get("Details", "")))

                    if blank_row is not None:
                        df_index = blank_row - 2
                        work = st.session_state["_simple_checker_df"]

                        if df_index in work.index:
                            preview_cols = [entity_col] + list(period_columns[:5])
                            preview_cols = [
                                c for c in preview_cols if c in work.columns
                            ]
                            st.caption("Row preview")
                            st.dataframe(
                                pd.DataFrame([
                                    work.loc[df_index, preview_cols]
                                ]),
                                use_container_width=True,
                                hide_index=True
                            )

                            new_name = st.text_input(
                                "Enter Entity Name",
                                key=f"blank_name_{blank_row}"
                            ).strip()

                            bc1, bc2, bc3 = st.columns(3)

                            if bc1.button(
                                "Apply Entity Name",
                                key=f"blank_apply_{blank_row}",
                                type="primary",
                                use_container_width=True
                            ):
                                if new_name:
                                    corrected = work.copy()
                                    corrected.loc[
                                        df_index, entity_col
                                    ] = new_name
                                    st.session_state[
                                        "_simple_checker_df"
                                    ] = corrected
                                    st.session_state[
                                        "_country_review_decisions"
                                    ][blank_pid] = "Fixed"
                                    st.session_state[
                                        "_simple_checker_log"
                                    ].append({
                                        "Entity": new_name,
                                        "Problem": "Blank Entity Name",
                                        "Method": "Manual Entity Name Entry",
                                        "Excel Row": blank_row
                                    })
                                    st.rerun()
                                else:
                                    st.warning(
                                        "Enter an entity name first."
                                    )

                            if bc2.button(
                                "Delete Row",
                                key=f"blank_delete_{blank_row}",
                                use_container_width=True
                            ):
                                corrected = work.drop(
                                    index=df_index
                                ).reset_index(drop=True)
                                st.session_state[
                                    "_simple_checker_df"
                                ] = corrected
                                st.session_state[
                                    "_country_review_decisions"
                                ][blank_pid] = "Fixed"
                                st.session_state[
                                    "_simple_checker_log"
                                ].append({
                                    "Entity": blank_label,
                                    "Problem": "Blank Entity Name",
                                    "Method": "Deleted Row",
                                    "Excel Row": blank_row
                                })
                                st.rerun()

                            if bc3.button(
                                "Ignore / Keep Original",
                                key=f"blank_ignore_{blank_row}",
                                use_container_width=True
                            ):
                                st.session_state[
                                    "_country_review_decisions"
                                ][blank_pid] = "Ignored"
                                st.rerun()

    if other_countries:
        with st.expander(
            f"Other Entities With Problems ({len(other_countries)} entities)",
            expanded=False
        ):
            st.caption(
                f"These entities never entered Top {int(important_rank)}, "
                "but they can still be fixed and reviewed here."
            )

            for country_name in other_countries:
                findings = country_audit[
                    country_audit["Entity"].astype(str).eq(country_name)
                ].reset_index(drop=True)

                best_rank = best_rank_by_entity.get(country_name)
                rank_text = f" • Best Rank #{best_rank}" if best_rank is not None else ""

                # Combine multiple Duplicate Row messages into one clear item.
                duplicate_rows = findings[
                    findings["Category"].astype(str).eq("Duplicate Row")
                ].copy()
                display_findings = findings[
                    ~findings["Category"].astype(str).eq("Duplicate Row")
                ].copy()

                if not duplicate_rows.empty:
                    combined = duplicate_rows.iloc[0].copy()
                    excel_rows = []
                    for detail in duplicate_rows["Details"].astype(str):
                        import re
                        match = re.search(r"Excel row\s+(\d+)", detail)
                        if match:
                            excel_rows.append(int(match.group(1)))
                    excel_rows = sorted(set(excel_rows))
                    combined["Details"] = (
                        "Identical duplicate rows detected: "
                        + ", ".join(f"Excel row {r}" for r in excel_rows)
                        + "."
                    )
                    combined["Suggestion"] = (
                        "If the rows are truly identical, keep one and remove the extra copy."
                    )
                    display_findings = pd.concat(
                        [display_findings, pd.DataFrame([combined])],
                        ignore_index=True
                    )

                with st.expander(
                    f"{country_name} — {len(display_findings)} problem(s){rank_text}",
                    expanded=False
                ):
                    for problem_number, finding in display_findings.reset_index(drop=True).iterrows():
                        category = str(finding["Category"])
                        period_text = str(finding.get("Period", "")).strip()
                        pid = _problem_id(country_name, finding)
                        decision = st.session_state["_country_review_decisions"].get(
                            pid, "Pending"
                        )

                        st.markdown(f"### {problem_number + 1}. {category}")
                        if period_text:
                            st.write(f"**Period:** {period_text}")

                        context = _context_rows(country_name, period_text)
                        if context:
                            st.dataframe(
                                pd.DataFrame(context),
                                use_container_width=True,
                                hide_index=True
                            )

                        details = str(finding.get("Details", "")).strip()
                        suggestion = str(finding.get("Suggestion", "")).strip()
                        if details:
                            st.write(details)
                        if suggestion:
                            st.caption(suggestion)

                        plan = None
                        interpolation_plan = None
                        if category == "Internal Gap":
                            plan = _series_fill_plan(country_name, period_text)
                        elif category in {
                            "Repeated Consecutive Value",
                            "Zero Run",
                            "Non-Numeric Value",
                            "Negative Value"
                        }:
                            interpolation_plan = _range_interpolation_plan(
                                country_name, period_text
                            )

                        if plan is not None and decision == "Pending":
                            st.write(
                                f"**Before:** {plan['before']:,.0f}   |   "
                                f"**After:** {plan['after']:,.0f}"
                            )
                            st.dataframe(
                                pd.DataFrame({
                                    "Period": [str(c) for c in plan["columns"]],
                                    "Proposed Value": plan["values"]
                                }),
                                use_container_width=True,
                                hide_index=True
                            )

                        if interpolation_plan is not None and decision == "Pending":
                            st.warning(
                                "Suggested Series Fill is available because valid "
                                "values exist before and after this problem range."
                            )
                            st.dataframe(
                                pd.DataFrame({
                                    "Period": [
                                        str(c) for c in interpolation_plan["columns"]
                                    ],
                                    "Proposed Value": interpolation_plan["values"]
                                }),
                                use_container_width=True,
                                hide_index=True
                            )

                        b1, b2, b3 = st.columns(3)

                        if decision == "Pending":
                            if plan is not None:
                                if b1.button(
                                    "Fix — Series Fill",
                                    key=f"lower_series_{abs(hash(pid))}",
                                    use_container_width=True
                                ):
                                    corrected = st.session_state["_simple_checker_df"].copy()
                                    for col, value in zip(
                                        plan["columns"], plan["values"]
                                    ):
                                        corrected.loc[plan["row_index"], col] = value
                                    st.session_state["_simple_checker_df"] = corrected
                                    st.session_state["_country_review_decisions"][pid] = "Fixed"
                                    st.session_state["_simple_checker_log"].append({
                                        "Entity": country_name,
                                        "Problem": category,
                                        "Period": period_text,
                                        "Method": "Series Fill",
                                        "Updated Cells": len(plan["columns"])
                                    })
                                    st.rerun()

                            elif interpolation_plan is not None:
                                if b1.button(
                                    "Apply Suggested Series Fill",
                                    key=f"lower_interp_{abs(hash(pid))}",
                                    use_container_width=True
                                ):
                                    corrected = st.session_state["_simple_checker_df"].copy()
                                    for col, value in zip(
                                        interpolation_plan["columns"],
                                        interpolation_plan["values"]
                                    ):
                                        corrected.loc[
                                            interpolation_plan["row_index"], col
                                        ] = value
                                    st.session_state["_simple_checker_df"] = corrected
                                    st.session_state["_country_review_decisions"][pid] = "Fixed"
                                    st.session_state["_simple_checker_log"].append({
                                        "Entity": country_name,
                                        "Problem": category,
                                        "Period": period_text,
                                        "Method": "Suggested Series Fill"
                                    })
                                    st.rerun()

                            elif category == "Duplicate Row" and len(duplicate_rows) >= 2:
                                duplicate_entity_name = str(country_name)
                                if b1.button(
                                    "Remove Duplicate Row",
                                    key=f"remove_dup_{normalize_name(duplicate_entity_name)}",
                                    type="primary",
                                    use_container_width=True
                                ):
                                    work = st.session_state["_simple_checker_df"].copy()
                                    entity_matches = work[
                                        work[entity_column].astype(str).str.strip().eq(
                                            duplicate_entity_name.strip()
                                        )
                                    ]
                                    if len(entity_matches) >= 2:
                                        first_idx = entity_matches.index[0]
                                        duplicate_indices = []
                                        first_row = work.loc[first_idx]
                                        for idx in entity_matches.index[1:]:
                                            if work.loc[idx].equals(first_row):
                                                duplicate_indices.append(idx)
                                        if duplicate_indices:
                                            work = work.drop(
                                                index=duplicate_indices
                                            ).reset_index(drop=True)
                                            st.session_state["_simple_checker_df"] = work
                                            st.session_state["_country_review_decisions"][pid] = "Fixed"
                                            st.session_state["_simple_checker_log"].append({
                                                "Entity": duplicate_entity_name,
                                                "Problem": "Duplicate Row",
                                                "Method": "Removed Identical Duplicate Row",
                                                "Rows Removed": len(duplicate_indices)
                                            })
                                            st.rerun()
                                    st.warning(
                                        "The rows are not identical, so the app did not remove anything."
                                    )

                            else:
                                manual_cols = _affected_period_columns(period_text)
                                if manual_cols:
                                    if b1.button(
                                        "Manual Entry",
                                        key=f"lower_manual_open_{abs(hash(pid))}",
                                        use_container_width=True
                                    ):
                                        st.session_state[
                                            f"_lower_manual_open_{abs(hash(pid))}"
                                        ] = True
                                else:
                                    b1.button(
                                        "Manual Review Required",
                                        key=f"lower_manual_disabled_{abs(hash(pid))}",
                                        disabled=True,
                                        use_container_width=True
                                    )

                            if b2.button(
                                "Ignore / Keep Original",
                                key=f"lower_ignore_{abs(hash(pid))}",
                                use_container_width=True
                            ):
                                st.session_state["_country_review_decisions"][pid] = "Ignored"
                                st.rerun()

                        else:
                            b1.success(
                                "Reviewed — Kept Original"
                                if decision == "Ignored"
                                else "Fixed"
                            )
                            if b2.button(
                                "Undo",
                                key=f"lower_undo_{abs(hash(pid))}",
                                use_container_width=True
                            ):
                                st.session_state["_country_review_decisions"][pid] = "Pending"
                                st.rerun()

                        manual_key = f"_lower_manual_open_{abs(hash(pid))}"
                        if decision == "Pending" and st.session_state.get(
                            manual_key, False
                        ):
                            manual_cols = _affected_period_columns(period_text)
                            row_i = _entity_row(country_name)
                            st.write("**Manual Entry**")
                            st.caption(
                                "Enter the value(s) you verified from your source."
                            )
                            manual_values = {}
                            for col in manual_cols:
                                current_raw = st.session_state[
                                    "_simple_checker_df"
                                ].loc[row_i, col]
                                current_num = pd.to_numeric(
                                    pd.Series([current_raw]), errors="coerce"
                                ).iloc[0]
                                default_value = (
                                    float(current_num)
                                    if not pd.isna(current_num)
                                    else 0.0
                                )
                                manual_values[col] = st.number_input(
                                    f"{col}",
                                    value=default_value,
                                    key=f"lower_manual_value_{abs(hash(pid))}_{str(col)}"
                                )

                            lm1, lm2 = st.columns(2)
                            if lm1.button(
                                "Apply Manual Entry",
                                key=f"lower_manual_apply_{abs(hash(pid))}",
                                type="primary",
                                use_container_width=True
                            ):
                                corrected = st.session_state[
                                    "_simple_checker_df"
                                ].copy()
                                for col, value in manual_values.items():
                                    corrected.loc[row_i, col] = value
                                st.session_state["_simple_checker_df"] = corrected
                                st.session_state["_country_review_decisions"][pid] = "Fixed"
                                st.session_state["_simple_checker_log"].append({
                                    "Entity": country_name,
                                    "Problem": category,
                                    "Period": period_text,
                                    "Method": "Manual Entry",
                                    "New Value": str({
                                        str(k): v for k, v in manual_values.items()
                                    })
                                })
                                st.session_state[manual_key] = False
                                st.rerun()

                            if lm2.button(
                                "Cancel Manual Entry",
                                key=f"lower_manual_cancel_{abs(hash(pid))}",
                                use_container_width=True
                            ):
                                st.session_state[manual_key] = False
                                st.rerun()

                        st.divider()

else:
    st.success("No entity problems found.")

if 'flag_check_status' in locals():
    st.caption(flag_check_status)


# =========================================================
# VIEW ALL PROBLEMS
# =========================================================

with st.expander(
    f"View All Problems ({len(audit_df)})",
    expanded=True
):
    if audit_df.empty:
        st.success("No problems found.")
    else:
        filter_options = ["All"] + sorted(
            audit_df["Category"].dropna().unique().tolist()
        )

        problem_filter = st.selectbox(
            "Show",
            filter_options
        )

        if problem_filter == "All":
            display_df = audit_df.copy()
        else:
            display_df = audit_df[
                audit_df["Category"].eq(problem_filter)
            ].copy()

        columns_to_show = [
            "Priority",
            "Severity",
            "Category",
            "Entity",
            "Best Rank",
            "Period",
            "Value",
            "Details",
            "Suggestion"
        ]

        st.dataframe(
            display_df[columns_to_show],
            use_container_width=True,
            hide_index=True
        )


with st.expander(
    "Completeness",
    expanded=False
):
    st.dataframe(
        completeness_df,
        use_container_width=True,
        hide_index=True
    )


# =========================================================
# STEP 5 — DOWNLOAD
# =========================================================

st.header("5. Finish")

correction_log = st.session_state.get(
    "_simple_checker_log",
    []
)

if correction_log:
    st.success(
        f"{len(correction_log)} correction(s) applied."
    )

    st.dataframe(
        pd.DataFrame(correction_log),
        use_container_width=True,
        hide_index=True
    )

corrected_workbook_bytes = make_corrected_workbook(
    uploaded_bytes,
    file_name,
    sheet_name,
    st.session_state["_simple_checker_df"]
)

corrected_name = (
    Path(file_name).stem
    + "_CORRECTED.xlsx"
)

d1, d2 = st.columns(2)

with d1:
    st.download_button(
        "Download Corrected Excel",
        data=corrected_workbook_bytes,
        file_name=corrected_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

with d2:
    report_output = io.BytesIO()

    with pd.ExcelWriter(
        report_output,
        engine="openpyxl"
    ) as writer:
        audit_df.to_excel(
            writer,
            sheet_name="Audit_Findings",
            index=False
        )

        completeness_df.to_excel(
            writer,
            sheet_name="Completeness",
            index=False
        )

        pd.DataFrame([{
            "File": file_name,
            "Sheet": sheet_name,
            "Entity Column": entity_column,
            "Rows": len(df),
            "Time Type": time_type,
            "Period Columns": len(period_columns),
            "Total Findings": len(audit_df),
            "Errors": int(
                audit_df["Severity"].eq("Error").sum()
                if not audit_df.empty else 0
            ),
            "Warnings": int(
                audit_df["Severity"].eq("Warning").sum()
                if not audit_df.empty else 0
            ),
            "Review Items": int(
                audit_df["Severity"].eq("Review").sum()
                if not audit_df.empty else 0
            )
        }]).to_excel(
            writer,
            sheet_name="Summary",
            index=False
        )

    report_output.seek(0)

    st.download_button(
        "Download Audit Report",
        data=report_output.getvalue(),
        file_name="Data_Rank_Hub_Excel_Check_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

if st.button(
    "Reset Corrections",
    use_container_width=True
):
    st.session_state["_simple_checker_df"] = df_source.copy()
    st.session_state["_simple_checker_log"] = []
    st.session_state["_country_review_decisions"] = {}
    st.session_state["_completed_countries"] = []
    st.session_state["_simple_checker_checked"] = False
    st.rerun()
