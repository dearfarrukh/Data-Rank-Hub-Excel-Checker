
import os
import re
import io
import hashlib
from difflib import get_close_matches
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# =========================================================
# DATA RANK HUB — SIMPLE EXCEL CHECKER
# Upload → Check → Fix → Download
# Same app can run locally or from Google Drive / Colab.
# =========================================================

st.set_page_config(
    page_title="Data Rank Hub Excel Checker",
    layout="wide"
)

st.title("Data Rank Hub Excel Checker")
st.caption("Upload → Check → Fix → Download")

APP_FOLDER = os.path.dirname(os.path.abspath(__file__))

# Keep flag folders relative to the app so the same code works
# on the home PC and inside a Google Drive / Colab app folder.
COUNTRY_FLAG_LIBRARY = os.path.join(
    APP_FOLDER, "flags", "flag_library"
)
STATE_FLAG_LIBRARY = os.path.join(
    APP_FOLDER, "states"
)

IMAGE_EXTENSIONS = (".png", ".jpg", ".jpeg")

ENTITY_CANDIDATES = [
    "Country", "State", "Entity", "City", "Company", "Name"
]

MONTH_NAMES = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12
}


# =========================================================
# HELPERS
# =========================================================

def detect_entity_column(df):
    for candidate in ENTITY_CANDIDATES:
        if candidate in df.columns:
            return candidate
    return df.columns[0]


def normalize_name(text):
    text = str(text).strip().lower()
    replacements = {
        "&": "and",
        "_": " ",
        "-": " ",
        ".": "",
        "'": "",
        "’": ""
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return " ".join(text.split())


def parse_period_column(column):
    text = str(column).strip()

    # Annual
    if re.fullmatch(r"\d{4}", text):
        year = int(text)
        if 1 <= year <= 2200:
            return {
                "type": "Annual",
                "year": year,
                "month": 1,
                "quarter": 1,
                "sort_key": year,
                "original_column": column,
                "display": text
            }

    # Monthly MM-YYYY / MM / YYYY
    match = re.fullmatch(r"(\d{1,2})\s*[-/]\s*(\d{4})", text)
    if match:
        month = int(match.group(1))
        year = int(match.group(2))
        if 1 <= month <= 12:
            return {
                "type": "Monthly",
                "year": year,
                "month": month,
                "quarter": ((month - 1) // 3) + 1,
                "sort_key": year * 12 + month,
                "original_column": column,
                "display": text
            }

    # Monthly YYYY-MM
    match = re.fullmatch(r"(\d{4})\s*[-/]\s*(\d{1,2})", text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return {
                "type": "Monthly",
                "year": year,
                "month": month,
                "quarter": ((month - 1) // 3) + 1,
                "sort_key": year * 12 + month,
                "original_column": column,
                "display": text
            }

    # Monthly Jan 2024
    match = re.fullmatch(r"([A-Za-z]+)\s+(\d{4})", text)
    if match:
        month_name = match.group(1).lower()
        year = int(match.group(2))
        if month_name in MONTH_NAMES:
            month = MONTH_NAMES[month_name]
            return {
                "type": "Monthly",
                "year": year,
                "month": month,
                "quarter": ((month - 1) // 3) + 1,
                "sort_key": year * 12 + month,
                "original_column": column,
                "display": text
            }

    # Quarterly 2024 Q1
    match = re.fullmatch(r"(\d{4})\s*[- ]?Q([1-4])", text, re.IGNORECASE)
    if match:
        year = int(match.group(1))
        quarter = int(match.group(2))
        return {
            "type": "Quarterly",
            "year": year,
            "month": (quarter - 1) * 3 + 1,
            "quarter": quarter,
            "sort_key": year * 4 + quarter,
            "original_column": column,
            "display": text
        }

    # Quarterly Q1 2024
    match = re.fullmatch(r"Q([1-4])\s*[- ]?(\d{4})", text, re.IGNORECASE)
    if match:
        quarter = int(match.group(1))
        year = int(match.group(2))
        return {
            "type": "Quarterly",
            "year": year,
            "month": (quarter - 1) * 3 + 1,
            "quarter": quarter,
            "sort_key": year * 4 + quarter,
            "original_column": column,
            "display": text
        }

    return None


def detect_period_columns(df):
    detected = []
    for column in df.columns:
        parsed = parse_period_column(column)
        if parsed is not None:
            detected.append(parsed)

    if not detected:
        return [], None

    counts = {"Annual": 0, "Monthly": 0, "Quarterly": 0}
    for item in detected:
        counts[item["type"]] += 1

    dominant_type = max(counts, key=counts.get)
    filtered = [
        item for item in detected
        if item["type"] == dominant_type
    ]
    return filtered, dominant_type


def is_true_blank(value):
    if value is None:
        return True
    if pd.isna(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def expected_period_keys(periods, time_type):
    if not periods:
        return []

    sorted_periods = sorted(periods, key=lambda x: x["sort_key"])
    first = sorted_periods[0]
    last = sorted_periods[-1]
    expected = []

    if time_type == "Annual":
        for year in range(first["year"], last["year"] + 1):
            expected.append({"key": year, "label": str(year)})

    elif time_type == "Monthly":
        start_index = first["year"] * 12 + first["month"]
        end_index = last["year"] * 12 + last["month"]

        for index in range(start_index, end_index + 1):
            year = (index - 1) // 12
            month = ((index - 1) % 12) + 1
            expected.append(
                {"key": index, "label": f"{month:02d} - {year}"}
            )

    else:
        start_index = first["year"] * 4 + first["quarter"]
        end_index = last["year"] * 4 + last["quarter"]

        for index in range(start_index, end_index + 1):
            year = (index - 1) // 4
            quarter = ((index - 1) % 4) + 1
            expected.append(
                {"key": index, "label": f"{year} Q{quarter}"}
            )

    return expected


def repeated_runs_for_row(row, period_columns, threshold):
    values = []
    for column in period_columns:
        numeric = pd.to_numeric(
            pd.Series([row[column]]),
            errors="coerce"
        ).iloc[0]
        values.append(numeric)

    results = []
    i = 0

    while i < len(values):
        if pd.isna(values[i]):
            i += 1
            continue

        j = i + 1
        while (
            j < len(values)
            and not pd.isna(values[j])
            and values[j] == values[i]
        ):
            j += 1

        run_length = j - i

        if run_length >= threshold:
            results.append({
                "start_index": i,
                "end_index": j - 1,
                "length": run_length,
                "value": values[i]
            })

        i = j

    return results


def compress_missing_ranges(missing_indexes, period_columns):
    if not missing_indexes:
        return []

    ranges = []
    start = previous = missing_indexes[0]

    for current in missing_indexes[1:]:
        if current == previous + 1:
            previous = current
            continue

        ranges.append((start, previous))
        start = previous = current

    ranges.append((start, previous))

    output = []

    for start_index, end_index in ranges:
        start_label = str(period_columns[start_index])
        end_label = str(period_columns[end_index])

        output.append({
            "start_index": start_index,
            "end_index": end_index,
            "label": (
                start_label
                if start_index == end_index
                else f"{start_label} → {end_label}"
            ),
            "length": end_index - start_index + 1
        })

    return output


def get_flag_names(folder):
    names = []
    if not os.path.isdir(folder):
        return names

    for filename in os.listdir(folder):
        if filename.lower().endswith(IMAGE_EXTENSIONS):
            names.append(os.path.splitext(filename)[0])

    return sorted(set(names))


def find_flag_exact(entity, folder):
    if not os.path.isdir(folder):
        return None

    target = normalize_name(entity)

    for filename in os.listdir(folder):
        if not filename.lower().endswith(IMAGE_EXTENSIONS):
            continue

        stem = os.path.splitext(filename)[0]
        if normalize_name(stem) == target:
            return stem

    return None


def suggest_flag_match(entity, available_names):
    normalized_to_original = {
        normalize_name(name): name
        for name in available_names
    }

    close = get_close_matches(
        normalize_name(entity),
        list(normalized_to_original.keys()),
        n=1,
        cutoff=0.72
    )

    if close:
        return normalized_to_original[close[0]]

    return ""


def detect_correction_decimals(dataframe, period_columns):
    values = []

    for column in period_columns:
        numeric = pd.to_numeric(
            dataframe[column],
            errors="coerce"
        ).dropna()

        if not numeric.empty:
            values.extend(numeric.astype(float).tolist())

    if not values:
        return 2

    sample = values[:5000]

    for decimals in range(0, 7):
        matching = 0

        for value in sample:
            if abs(value - round(value, decimals)) < 10 ** (-(decimals + 4)):
                matching += 1

        if matching / len(sample) >= 0.95:
            return decimals

    return 6


def surrounding_values(df, row_index, start_index, end_index, period_columns):
    previous_value = None
    next_value = None

    for i in range(start_index - 1, -1, -1):
        value = pd.to_numeric(
            pd.Series([df.loc[row_index, period_columns[i]]]),
            errors="coerce"
        ).iloc[0]

        if not pd.isna(value):
            previous_value = float(value)
            break

    for i in range(end_index + 1, len(period_columns)):
        value = pd.to_numeric(
            pd.Series([df.loc[row_index, period_columns[i]]]),
            errors="coerce"
        ).iloc[0]

        if not pd.isna(value):
            next_value = float(value)
            break

    return previous_value, next_value


def linear_fill_values(count, previous_value, next_value, decimals):
    if (
        count <= 0
        or previous_value is None
        or next_value is None
    ):
        return None

    step = (next_value - previous_value) / (count + 1)

    return [
        round(previous_value + step * (i + 1), decimals)
        for i in range(count)
    ]


def make_corrected_workbook(
    uploaded_bytes,
    original_file_name,
    selected_sheet_name,
    corrected_df
):
    output = io.BytesIO()

    if original_file_name.lower().endswith(".csv"):
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            corrected_df.to_excel(
                writer,
                sheet_name="Corrected_Data",
                index=False
            )
    else:
        workbook = load_workbook(io.BytesIO(uploaded_bytes))
        worksheet = workbook[selected_sheet_name]

        for col_index, column_name in enumerate(
            corrected_df.columns,
            start=1
        ):
            worksheet.cell(
                row=1,
                column=col_index
            ).value = str(column_name)

        for row_offset, (_, data_row) in enumerate(
            corrected_df.iterrows(),
            start=2
        ):
            for col_index, column_name in enumerate(
                corrected_df.columns,
                start=1
            ):
                value = data_row[column_name]

                if pd.isna(value):
                    excel_value = None
                elif hasattr(value, "item"):
                    excel_value = value.item()
                else:
                    excel_value = value

                worksheet.cell(
                    row=row_offset,
                    column=col_index
                ).value = excel_value

        workbook.save(output)

    output.seek(0)
    return output.getvalue()


# =========================================================
# STEP 1 — UPLOAD
# =========================================================

st.header("1. Upload Excel")

uploaded_file = st.file_uploader(
    "Choose Excel or CSV file",
    type=["xlsx", "xls", "csv"]
)

if uploaded_file is None:
    st.info("Upload a file to begin.")
    st.stop()

uploaded_bytes = uploaded_file.getvalue()
file_name = uploaded_file.name

if file_name.lower().endswith(".csv"):
    df_source = pd.read_csv(
        io.BytesIO(uploaded_bytes),
        keep_default_na=False
    )
    sheet_name = "CSV"
else:
    excel_book = pd.ExcelFile(io.BytesIO(uploaded_bytes))
    if len(excel_book.sheet_names) == 1:
        sheet_name = excel_book.sheet_names[0]
        st.caption(f"Sheet: {sheet_name}")
    else:
        sheet_name = st.selectbox(
            "Sheet",
            excel_book.sheet_names
        )

    df_source = pd.read_excel(
        io.BytesIO(uploaded_bytes),
        sheet_name=sheet_name,
        keep_default_na=False
    )

detected_entity = detect_entity_column(df_source)
periods, time_type = detect_period_columns(df_source)

if not periods:
    st.error("No annual, monthly, or quarterly columns were detected.")
    st.stop()

entity_column = detected_entity

period_columns = [
    item["original_column"]
    for item in periods
]

file_key = hashlib.md5(
    (
        file_name
        + "|"
        + str(sheet_name)
        + "|"
        + str(len(uploaded_bytes))
    ).encode("utf-8")
).hexdigest()

if st.session_state.get("_simple_checker_file_key") != file_key:
    st.session_state["_simple_checker_file_key"] = file_key
    st.session_state["_simple_checker_df"] = df_source.copy()
    st.session_state["_simple_checker_checked"] = False
    st.session_state["_simple_checker_log"] = []

working_df = st.session_state["_simple_checker_df"].copy()

for column in period_columns:
    working_df[column] = working_df[column].astype("object")

st.session_state["_simple_checker_df"] = working_df

c1, c2, c3, c4 = st.columns(4)

c1.metric("Entities", len(working_df))
c2.metric("Periods", len(period_columns))
c3.metric("Time Type", time_type)
c4.metric(
    "Range",
    f"{period_columns[0]} → {period_columns[-1]}"
)

st.caption(
    f"Entity column detected automatically: {entity_column}"
)


# =========================================================
# ADVANCED SETTINGS
# =========================================================

with st.expander("Advanced Settings", expanded=False):

    a1, a2 = st.columns(2)

    with a1:
        data_type = st.selectbox(
            "Data Type",
            [
                "Regular Data",
                "Cumulative Totals"
            ],
            help=(
                "Use Cumulative Totals for data such as total COVID deaths. "
                "Repeated totals are allowed, but decreases are flagged."
            )
        )

        repeated_threshold = st.number_input(
            "Repeated Value Threshold",
            min_value=2,
            max_value=20,
            value=2,
            step=1
        )

        important_rank = st.number_input(
            "Important Rank",
            min_value=1,
            max_value=max(1, len(working_df)),
            value=min(20, len(working_df)),
            step=1
        )

    with a2:
        suspicious_jump_threshold = st.number_input(
            "Suspicious Jump %",
            min_value=100,
            max_value=10000,
            value=500,
            step=50
        )

        check_negative_values = st.checkbox(
            "Check Negative Values",
            value=True
        )

        flag_library_mode = st.selectbox(
            "Flag Check",
            [
                "Auto",
                "Countries / Historical",
                "US States",
                "Do Not Check Flags"
            ]
        )


# =========================================================
# STEP 2 — CHECK
# =========================================================

st.header("2. Check File")

check_clicked = st.button(
    "CHECK EXCEL",
    type="primary",
    use_container_width=True
)

if check_clicked:
    st.session_state["_simple_checker_checked"] = True

if not st.session_state.get("_simple_checker_checked", False):
    st.stop()

df = st.session_state["_simple_checker_df"].copy()
correction_decimals = detect_correction_decimals(
    df_source,
    period_columns
)

audit_rows = []

def add_audit(
    category,
    severity,
    entity="",
    period="",
    value="",
    details="",
    suggestion="",
    fixable=False
):
    audit_rows.append({
        "Category": category,
        "Severity": severity,
        "Entity": entity,
        "Period": period,
        "Value": value,
        "Details": details,
        "Suggestion": suggestion,
        "Fixable": fixable
    })

progress = st.progress(0)
status = st.empty()

# ---------------------------------------------------------
# Structure
# ---------------------------------------------------------

status.write("Checking structure...")
progress.progress(10)

blank_entity_mask = (
    df[entity_column].isna()
    | df[entity_column].astype(str).str.strip().eq("")
)

for index in df.index[blank_entity_mask]:
    add_audit(
        "Blank Entity Name",
        "Error",
        details=f"Blank entity name on Excel row {index + 2}."
    )

entity_clean = df[entity_column].astype(str).str.strip()

duplicate_mask = (
    entity_clean.ne("")
    & entity_clean.duplicated(keep=False)
)

for entity_name, group in df[duplicate_mask].groupby(
    entity_clean[duplicate_mask]
):
    excel_rows = [int(index) + 2 for index in group.index]
    add_audit(
        "Duplicate Entity",
        "Error",
        entity=entity_name,
        details=f"Appears on Excel rows {excel_rows}."
    )

full_duplicate_mask = df.duplicated(keep=False)

for index in df.index[full_duplicate_mask]:
    add_audit(
        "Duplicate Row",
        "Error",
        entity=str(df.loc[index, entity_column]),
        details=f"Entire row duplicated. Excel row {index + 2}."
    )

# ---------------------------------------------------------
# Date / period
# ---------------------------------------------------------

status.write("Checking dates...")
progress.progress(22)

period_key_map = {}

for period in periods:
    period_key_map.setdefault(
        period["sort_key"],
        []
    ).append(
        str(period["original_column"])
    )

for labels in period_key_map.values():
    if len(labels) > 1:
        add_audit(
            "Duplicate Period",
            "Error",
            period=" / ".join(labels),
            details="The same period appears more than once."
        )

actual_keys = [item["sort_key"] for item in periods]

if actual_keys != sorted(actual_keys):
    add_audit(
        "Periods Out of Order",
        "Warning",
        details="Time columns are not arranged chronologically."
    )

expected = expected_period_keys(periods, time_type)
existing_keys = set(actual_keys)

for item in expected:
    if item["key"] not in existing_keys:
        add_audit(
            "Missing Period",
            "Warning",
            period=item["label"],
            details=f"{item['label']} is missing from the time-series columns."
        )

# ---------------------------------------------------------
# Missing / invalid
# ---------------------------------------------------------

status.write("Checking missing and invalid data...")
progress.progress(36)

for row_index, row in df.iterrows():
    entity = str(row[entity_column]).strip()
    converted = []

    for column in period_columns:
        raw_value = row[column]

        if is_true_blank(raw_value):
            converted.append(("blank", None, raw_value))
            continue

        numeric_value = pd.to_numeric(
            pd.Series([raw_value]),
            errors="coerce"
        ).iloc[0]

        if pd.isna(numeric_value):
            converted.append(("invalid", None, raw_value))
        else:
            converted.append(("number", float(numeric_value), raw_value))

    valid_positions = [
        i for i, item in enumerate(converted)
        if item[0] == "number"
    ]

    for i, item in enumerate(converted):
        kind, numeric_value, raw_value = item

        if kind == "invalid":
            add_audit(
                "Non-Numeric Value",
                "Error",
                entity=entity,
                period=str(period_columns[i]),
                value=str(raw_value),
                details=f"Expected a number on Excel row {row_index + 2}.",
                suggestion="Replace with the correct number."
            )

    if valid_positions:
        first_valid = valid_positions[0]
        last_valid = valid_positions[-1]

        missing_inside = [
            i for i in range(first_valid + 1, last_valid)
            if converted[i][0] == "blank"
        ]

        for missing_range in compress_missing_ranges(
            missing_inside,
            period_columns
        ):
            add_audit(
                "Internal Gap",
                "Warning",
                entity=entity,
                period=missing_range["label"],
                details=(
                    f"{missing_range['length']} missing period(s) "
                    f"inside the series. Excel row {row_index + 2}."
                ),
                suggestion="Series Fill is available when there is a real value before and after.",
                fixable=True
            )

# ---------------------------------------------------------
# Number checks
# ---------------------------------------------------------

status.write("Checking numbers...")
progress.progress(52)

for row_index, row in df.iterrows():
    entity = str(row[entity_column]).strip()

    runs = repeated_runs_for_row(
        row,
        period_columns,
        int(repeated_threshold)
    )

    for run in runs:
        start_column = str(period_columns[run["start_index"]])
        end_column = str(period_columns[run["end_index"]])
        value = run["value"]

        if float(value) == 0:
            add_audit(
                "Zero Run",
                "Review",
                entity=entity,
                period=f"{start_column} → {end_column}",
                value=value,
                details=f"Zero appears in {run['length']} consecutive periods.",
                suggestion="Review whether zeros are real values or placeholders."
            )

        elif data_type == "Regular Data":
            add_audit(
                "Repeated Consecutive Value",
                "Review",
                entity=entity,
                period=f"{start_column} → {end_column}",
                value=value,
                details=f"Same non-zero value appears in {run['length']} consecutive periods.",
                suggestion="Review whether this is genuine or accidental forward-fill."
            )

    previous_numeric = None
    previous_column = None

    for column in period_columns:
        current_numeric = pd.to_numeric(
            pd.Series([row[column]]),
            errors="coerce"
        ).iloc[0]

        if pd.isna(current_numeric):
            continue

        current_numeric = float(current_numeric)

        if check_negative_values and current_numeric < 0:
            add_audit(
                "Negative Value",
                "Review",
                entity=entity,
                period=str(column),
                value=current_numeric,
                details=f"Negative value on Excel row {row_index + 2}.",
                suggestion="Review whether negatives are valid for this dataset."
            )

        if (
            previous_numeric is not None
            and previous_numeric != 0
        ):
            change_percent = abs(
                (
                    current_numeric - previous_numeric
                )
                / abs(previous_numeric)
                * 100
            )

            if change_percent > float(suspicious_jump_threshold):
                add_audit(
                    "Suspicious Jump",
                    "Review",
                    entity=entity,
                    period=f"{previous_column} → {column}",
                    value=current_numeric,
                    details=(
                        f"Changed by {change_percent:,.1f}% "
                        f"from {previous_numeric:,.6g} "
                        f"to {current_numeric:,.6g}."
                    ),
                    suggestion="Review this large change."
                )

        if (
            data_type == "Cumulative Totals"
            and previous_numeric is not None
            and current_numeric < previous_numeric
        ):
            add_audit(
                "Cumulative Value Decreased",
                "Error",
                entity=entity,
                period=f"{previous_column} → {column}",
                value=current_numeric,
                details=(
                    f"Cumulative value fell from "
                    f"{previous_numeric:,.6g} "
                    f"to {current_numeric:,.6g}."
                ),
                suggestion="Cumulative totals normally should not decrease."
            )

        previous_numeric = current_numeric
        previous_column = str(column)

# ---------------------------------------------------------
# Priority ranks
# ---------------------------------------------------------

status.write("Checking important entities...")
progress.progress(66)

best_rank_by_entity = {}

for period in periods:
    column = period["original_column"]
    numeric_values = pd.to_numeric(
        df[column],
        errors="coerce"
    )
    ranks = numeric_values.rank(
        method="min",
        ascending=False
    )

    for row_index, rank_value in ranks.items():
        if pd.isna(rank_value):
            continue

        entity_name = str(
            df.loc[row_index, entity_column]
        ).strip()

        if not entity_name:
            continue

        rank_number = int(rank_value)
        current_best = best_rank_by_entity.get(entity_name)

        if (
            current_best is None
            or rank_number < current_best
        ):
            best_rank_by_entity[entity_name] = rank_number

# Important start/end coverage
for row_index, row in df.iterrows():
    entity_name = str(row[entity_column]).strip()

    if not entity_name:
        continue

    best_rank = best_rank_by_entity.get(entity_name)

    if (
        best_rank is None
        or best_rank > int(important_rank)
    ):
        continue

    numeric_series = pd.to_numeric(
        row[period_columns],
        errors="coerce"
    )

    valid_positions = [
        i for i, value in enumerate(numeric_series.tolist())
        if not pd.isna(value)
    ]

    if not valid_positions:
        continue

    leading_missing = list(range(0, valid_positions[0]))
    trailing_missing = list(
        range(valid_positions[-1] + 1, len(period_columns))
    )

    for missing_range in (
        compress_missing_ranges(leading_missing, period_columns)
        + compress_missing_ranges(trailing_missing, period_columns)
    ):
        add_audit(
            "Important Coverage Gap",
            "Warning",
            entity=entity_name,
            period=missing_range["label"],
            details=(
                f"{missing_range['length']} missing period(s) "
                f"for an entity that reached Top {int(important_rank)}."
            ),
            suggestion="Review whether reliable data can be added."
        )

# ---------------------------------------------------------
# Flags
# ---------------------------------------------------------

status.write("Checking flags...")
progress.progress(78)

flag_library_available = (
    os.path.isdir(COUNTRY_FLAG_LIBRARY)
    or os.path.isdir(STATE_FLAG_LIBRARY)
)

if flag_library_mode == "Do Not Check Flags":
    flag_check_status = "Flag checking turned off."

elif not flag_library_available:
    flag_check_status = (
        "Flag library is not connected on this online app, "
        "so flag checking was skipped."
    )

else:
    country_names = get_flag_names(COUNTRY_FLAG_LIBRARY)
    state_names = get_flag_names(STATE_FLAG_LIBRARY)

    entities = (
        df[entity_column]
        .dropna()
        .astype(str)
        .str.strip()
    )
    entities = entities[entities.ne("")]

    if flag_library_mode == "Countries / Historical":
        selected_folder = COUNTRY_FLAG_LIBRARY
        available_names = country_names
        selected_label = "Countries / Historical"

    elif flag_library_mode == "US States":
        selected_folder = STATE_FLAG_LIBRARY
        available_names = state_names
        selected_label = "US States"

    else:
        country_match_count = sum(
            find_flag_exact(entity, COUNTRY_FLAG_LIBRARY) is not None
            for entity in entities
        )
        state_match_count = sum(
            find_flag_exact(entity, STATE_FLAG_LIBRARY) is not None
            for entity in entities
        )

        if state_match_count > country_match_count:
            selected_folder = STATE_FLAG_LIBRARY
            available_names = state_names
            selected_label = "US States"
        else:
            selected_folder = COUNTRY_FLAG_LIBRARY
            available_names = country_names
            selected_label = "Countries / Historical"

    flag_check_status = f"Flag library used: {selected_label}"

    for entity in sorted(set(entities)):
        if find_flag_exact(entity, selected_folder) is not None:
            continue

        suggestion = suggest_flag_match(entity, available_names)

        if suggestion:
            add_audit(
                "Flag Name Mismatch",
                "Review",
                entity=entity,
                details=f"No exact flag filename matches '{entity}'.",
                suggestion=f"Possible match: {suggestion}"
            )
        else:
            add_audit(
                "Missing Flag",
                "Warning",
                entity=entity,
                details=f"No matching flag found in {selected_label} library."
            )

# ---------------------------------------------------------
# Completeness
# ---------------------------------------------------------

status.write("Calculating completeness...")
progress.progress(90)

completeness_rows = []

for _, row in df.iterrows():
    entity = str(row[entity_column]).strip()

    numeric_series = pd.to_numeric(
        row[period_columns],
        errors="coerce"
    )

    valid_count = int(numeric_series.notna().sum())
    missing_count = len(period_columns) - valid_count

    valid_positions = [
        i for i, value in enumerate(numeric_series.tolist())
        if not pd.isna(value)
    ]

    first_available = (
        str(period_columns[valid_positions[0]])
        if valid_positions
        else ""
    )

    last_available = (
        str(period_columns[valid_positions[-1]])
        if valid_positions
        else ""
    )

    completeness_rows.append({
        "Entity": entity,
        "Completeness %": round(
            valid_count / len(period_columns) * 100,
            1
        ) if period_columns else 0,
        "Observations": valid_count,
        "Missing": missing_count,
        "First Available": first_available,
        "Last Available": last_available
    })

# Priority fields
for finding in audit_rows:
    entity_name = str(finding.get("Entity", "")).strip()
    best_rank = best_rank_by_entity.get(entity_name)

    finding["Best Rank"] = (
        best_rank if best_rank is not None else ""
    )

    if (
        best_rank is not None
        and best_rank <= int(important_rank)
    ):
        finding["Priority"] = "HIGH PRIORITY"
    elif entity_name:
        finding["Priority"] = "Lower Priority"
    else:
        finding["Priority"] = "General"

audit_df = pd.DataFrame(audit_rows)
completeness_df = pd.DataFrame(completeness_rows)

if not audit_df.empty:
    order = {
        "HIGH PRIORITY": 0,
        "General": 1,
        "Lower Priority": 2
    }

    audit_df["_PriorityOrder"] = (
        audit_df["Priority"]
        .map(order)
        .fillna(9)
    )

    audit_df = (
        audit_df
        .sort_values(
            ["_PriorityOrder", "Best Rank", "Severity", "Category", "Entity"],
            na_position="last"
        )
        .drop(columns=["_PriorityOrder"])
        .reset_index(drop=True)
    )

progress.progress(100)
status.write("Check complete.")


# =========================================================
# STEP 3 — SIMPLE RESULT
# =========================================================

st.header("3. Problems")

if audit_df.empty:
    st.success("READY FOR BAR CHART RACE — No problems found.")
else:
    flag_categories = ["Missing Flag", "Flag Name Mismatch"]

    must_fix = int(
        (
            audit_df["Severity"].eq("Error")
            & ~audit_df["Category"].isin(flag_categories)
        ).sum()
    )

    should_review = int(
        (
            audit_df["Severity"].isin(["Warning", "Review"])
            & ~audit_df["Category"].isin(flag_categories)
        ).sum()
    )

    flag_problems = int(
        audit_df["Category"].isin(flag_categories).sum()
    )

    missing_problems = int(
        audit_df["Category"].isin(
            ["Internal Gap", "Important Coverage Gap", "Missing Period"]
        ).sum()
    )

    s1, s2, s3, s4 = st.columns(4)

    s1.metric("Must Fix", must_fix)
    s2.metric("Needs Review", should_review)
    s3.metric("Missing Data", missing_problems)
    s4.metric("Flag Problems", flag_problems)

    if must_fix > 0:
        st.error(
            f"{must_fix} problem(s) should be fixed before making the video."
        )
    elif should_review > 0:
        st.warning(
            "No critical data errors, but some items should be reviewed."
        )
    else:
        st.success("Data checks look good.")

# =========================================================
# STEP 4 — FIX
# =========================================================

st.header("4. Review Entities")

# Entity-first review state
if "_country_review_decisions" not in st.session_state:
    st.session_state["_country_review_decisions"] = {}
if "_completed_countries" not in st.session_state:
    st.session_state["_completed_countries"] = []

def _problem_id(entity_name, finding):
    return "||".join([
        str(entity_name),
        str(finding.get("Category", "")),
        str(finding.get("Period", "")),
        str(finding.get("Details", ""))
    ])

def _entity_row(entity_name):
    matches = df.index[
        df[entity_column].astype(str).str.strip().eq(str(entity_name).strip())
    ].tolist()
    return matches[0] if matches else None

def _context_rows(entity_name, period_text):
    """Show one value before and after the affected numerical period/range."""
    row_index = _entity_row(entity_name)
    if row_index is None:
        return []

    labels = [str(c) for c in period_columns]
    pieces = [p.strip() for p in str(period_text).split("→")]
    indexes = [labels.index(p) for p in pieces if p in labels]
    if not indexes:
        return []

    start = max(0, min(indexes) - 1)
    end = min(len(labels) - 1, max(indexes) + 1)
    rows = []

    for i in range(start, end + 1):
        col = period_columns[i]
        raw = st.session_state["_simple_checker_df"].loc[row_index, col]
        numeric = pd.to_numeric(pd.Series([raw]), errors="coerce").iloc[0]
        rows.append({
            "Period": str(col),
            "Value": "Missing" if pd.isna(numeric) else numeric
        })
    return rows

def _series_fill_plan(entity_name, period_text):
    row_index = _entity_row(entity_name)
    if row_index is None:
        return None

    labels = [str(c) for c in period_columns]
    if "→" in str(period_text):
        start_label, end_label = [p.strip() for p in str(period_text).split("→", 1)]
    else:
        start_label = end_label = str(period_text).strip()

    if start_label not in labels or end_label not in labels:
        return None

    start_index = labels.index(start_label)
    end_index = labels.index(end_label)
    prev_value, next_value = surrounding_values(
        st.session_state["_simple_checker_df"],
        row_index,
        start_index,
        end_index,
        period_columns
    )
    if prev_value is None or next_value is None:
        return None

    cols = period_columns[start_index:end_index + 1]
    values = linear_fill_values(
        len(cols), prev_value, next_value, int(correction_decimals)
    )
    if values is None:
        return None

    return {
        "row_index": row_index,
        "columns": cols,
        "values": values,
        "before": prev_value,
        "after": next_value
    }

if not audit_df.empty:
    country_audit = audit_df[
        audit_df["Entity"].astype(str).str.strip().ne("")
    ].copy()

    # Priority = any entity that entered selected Top N at least once.
    priority_countries = []
    other_countries = []
    for name in country_audit["Entity"].astype(str).drop_duplicates():
        rank = best_rank_by_entity.get(name)
        if rank is not None and rank <= int(important_rank):
            priority_countries.append(name)
        else:
            other_countries.append(name)

    priority_countries.sort(key=lambda x: best_rank_by_entity.get(x, 999999))
    other_countries.sort()

    completed_set = set(st.session_state["_completed_countries"])

    priority_remaining = [
        c for c in priority_countries
        if c not in completed_set
    ]
    priority_completed = [
        c for c in priority_countries
        if c in completed_set
    ]

    m1, m2, m3 = st.columns(3)
    m1.metric(f"Entities Ever in Top {int(important_rank)}", len(priority_countries))
    m2.metric("Priority Remaining", len(priority_remaining))
    m3.metric("Priority Completed", len(priority_completed))

    if priority_countries:
        st.info(
            f"Top {int(important_rank)} entities with detected problems: "
            + ", ".join(priority_countries)
        )

    st.subheader("Priority Entities Needing Review")

    for country_name in priority_remaining:
        findings = country_audit[
            country_audit["Entity"].astype(str).eq(country_name)
        ].reset_index(drop=True)

        best_rank = best_rank_by_entity.get(country_name)
        rank_text = f" • Best Rank #{best_rank}" if best_rank is not None else ""

        with st.expander(
            f"{country_name} — {len(findings)} problem(s){rank_text}",
            expanded=False
        ):
            for problem_number, finding in findings.iterrows():
                category = str(finding["Category"])
                period_text = str(finding.get("Period", "")).strip()
                pid = _problem_id(country_name, finding)
                decision = st.session_state["_country_review_decisions"].get(
                    pid, "Pending"
                )

                st.markdown(f"### {problem_number + 1}. {category}")
                if period_text:
                    st.write(f"**Period:** {period_text}")

                # For numerical problems, show the affected values plus one before/after.
                context = _context_rows(country_name, period_text)
                if context:
                    st.dataframe(
                        pd.DataFrame(context),
                        use_container_width=True,
                        hide_index=True
                    )

                details = str(finding.get("Details", "")).strip()
                suggestion = str(finding.get("Suggestion", "")).strip()
                if details:
                    st.write(details)
                if suggestion:
                    st.caption(suggestion)

                plan = None
                if category == "Internal Gap":
                    plan = _series_fill_plan(country_name, period_text)

                if plan is not None and decision == "Pending":
                    st.write(
                        f"**Before:** {plan['before']:,.6g}   |   "
                        f"**After:** {plan['after']:,.6g}"
                    )
                    st.write("**Proposed Series Fill:**")
                    st.dataframe(
                        pd.DataFrame({
                            "Period": [str(c) for c in plan["columns"]],
                            "Proposed Value": plan["values"]
                        }),
                        use_container_width=True,
                        hide_index=True
                    )

                c1, c2, c3 = st.columns(3)

                if decision == "Pending":
                    if plan is not None:
                        if c1.button(
                            "Fix — Series Fill",
                            key=f"country_fix_{abs(hash(pid))}",
                            use_container_width=True
                        ):
                            corrected = st.session_state["_simple_checker_df"].copy()
                            for col, value in zip(plan["columns"], plan["values"]):
                                corrected.loc[plan["row_index"], col] = value
                            st.session_state["_simple_checker_df"] = corrected
                            st.session_state["_country_review_decisions"][pid] = "Fixed"
                            st.session_state["_simple_checker_log"].append({
                                "Entity": country_name,
                                "Problem": category,
                                "Period": period_text,
                                "Method": "Series Fill",
                                "Updated Cells": len(plan["columns"])
                            })
                            st.rerun()
                    else:
                        c1.button(
                            "Fix",
                            key=f"manual_{abs(hash(pid))}",
                            disabled=True,
                            help="No safe automatic fix. Review the before/after values and correct manually if needed.",
                            use_container_width=True
                        )

                    if c2.button(
                        "Ignore / Keep Original",
                        key=f"country_ignore_{abs(hash(pid))}",
                        use_container_width=True
                    ):
                        st.session_state["_country_review_decisions"][pid] = "Ignored"
                        st.rerun()
                else:
                    c1.success(decision)
                    if c2.button(
                        "Undo",
                        key=f"country_undo_{abs(hash(pid))}",
                        use_container_width=True
                    ):
                        st.session_state["_country_review_decisions"][pid] = "Pending"
                        st.rerun()

                st.divider()

            pending = 0
            for _, finding in findings.iterrows():
                pid = _problem_id(country_name, finding)
                if st.session_state["_country_review_decisions"].get(
                    pid, "Pending"
                ) == "Pending":
                    pending += 1

            if pending == 0:
                if st.button(
                    f"ENTITY COMPLETE — {country_name}",
                    key=f"complete_{normalize_name(country_name)}",
                    type="primary",
                    use_container_width=True
                ):
                    if country_name not in st.session_state["_completed_countries"]:
                        st.session_state["_completed_countries"].append(country_name)
                    st.rerun()
            else:
                st.caption(
                    f"{pending} problem(s) remaining. Fix or Ignore each one to complete this entity."
                )

    if priority_completed:
        with st.expander(
            f"Completed Priority Entities ({len(priority_completed)})",
            expanded=False
        ):
            for country_name in priority_completed:
                cc1, cc2 = st.columns([4, 1])
                cc1.write(f"✅ {country_name}")
                if cc2.button(
                    "Reopen",
                    key=f"reopen_{normalize_name(country_name)}"
                ):
                    st.session_state["_completed_countries"].remove(country_name)
                    st.rerun()

    if other_countries:
        with st.expander(
            f"Lower Priority Problems ({len(other_countries)} entities)",
            expanded=False
        ):
            st.caption(
                f"These entities never entered Top {int(important_rank)}. "
                "They are kept separate so they do not slow down the priority review."
            )
            for country_name in other_countries:
                findings = country_audit[
                    country_audit["Entity"].astype(str).eq(country_name)
                ].reset_index(drop=True)
                best_rank = best_rank_by_entity.get(country_name)
                rank_text = f" • Best Rank #{best_rank}" if best_rank is not None else ""
                with st.expander(
                    f"{country_name} — {len(findings)} problem(s){rank_text}",
                    expanded=False
                ):
                    for problem_number, finding in findings.iterrows():
                        category = str(finding["Category"])
                        period_text = str(finding.get("Period", "")).strip()
                        st.markdown(f"**{problem_number + 1}. {category}**")
                        if period_text:
                            st.write(f"Period: **{period_text}**")
                        context = _context_rows(country_name, period_text)
                        if context:
                            st.dataframe(
                                pd.DataFrame(context),
                                use_container_width=True,
                                hide_index=True
                            )
                        details = str(finding.get("Details", "")).strip()
                        suggestion = str(finding.get("Suggestion", "")).strip()
                        if details:
                            st.write(details)
                        if suggestion:
                            st.caption(suggestion)
                        st.divider()
else:
    st.success("No entity problems found.")

if 'flag_check_status' in locals():
    st.caption(flag_check_status)


# =========================================================
# VIEW ALL PROBLEMS
# =========================================================

with st.expander(
    f"View All Problems ({len(audit_df)})",
    expanded=True
):
    if audit_df.empty:
        st.success("No problems found.")
    else:
        filter_options = ["All"] + sorted(
            audit_df["Category"].dropna().unique().tolist()
        )

        problem_filter = st.selectbox(
            "Show",
            filter_options
        )

        if problem_filter == "All":
            display_df = audit_df.copy()
        else:
            display_df = audit_df[
                audit_df["Category"].eq(problem_filter)
            ].copy()

        columns_to_show = [
            "Priority",
            "Severity",
            "Category",
            "Entity",
            "Best Rank",
            "Period",
            "Value",
            "Details",
            "Suggestion"
        ]

        st.dataframe(
            display_df[columns_to_show],
            use_container_width=True,
            hide_index=True
        )


with st.expander(
    "Completeness",
    expanded=False
):
    st.dataframe(
        completeness_df,
        use_container_width=True,
        hide_index=True
    )


# =========================================================
# STEP 5 — DOWNLOAD
# =========================================================

st.header("5. Finish")

correction_log = st.session_state.get(
    "_simple_checker_log",
    []
)

if correction_log:
    st.success(
        f"{len(correction_log)} correction(s) applied."
    )

    st.dataframe(
        pd.DataFrame(correction_log),
        use_container_width=True,
        hide_index=True
    )

corrected_workbook_bytes = make_corrected_workbook(
    uploaded_bytes,
    file_name,
    sheet_name,
    st.session_state["_simple_checker_df"]
)

corrected_name = (
    Path(file_name).stem
    + "_CORRECTED.xlsx"
)

d1, d2 = st.columns(2)

with d1:
    st.download_button(
        "Download Corrected Excel",
        data=corrected_workbook_bytes,
        file_name=corrected_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

with d2:
    report_output = io.BytesIO()

    with pd.ExcelWriter(
        report_output,
        engine="openpyxl"
    ) as writer:
        audit_df.to_excel(
            writer,
            sheet_name="Audit_Findings",
            index=False
        )

        completeness_df.to_excel(
            writer,
            sheet_name="Completeness",
            index=False
        )

        pd.DataFrame([{
            "File": file_name,
            "Sheet": sheet_name,
            "Entity Column": entity_column,
            "Rows": len(df),
            "Time Type": time_type,
            "Period Columns": len(period_columns),
            "Total Findings": len(audit_df),
            "Errors": int(
                audit_df["Severity"].eq("Error").sum()
                if not audit_df.empty else 0
            ),
            "Warnings": int(
                audit_df["Severity"].eq("Warning").sum()
                if not audit_df.empty else 0
            ),
            "Review Items": int(
                audit_df["Severity"].eq("Review").sum()
                if not audit_df.empty else 0
            )
        }]).to_excel(
            writer,
            sheet_name="Summary",
            index=False
        )

    report_output.seek(0)

    st.download_button(
        "Download Audit Report",
        data=report_output.getvalue(),
        file_name="Data_Rank_Hub_Excel_Check_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

if st.button(
    "Reset Corrections",
    use_container_width=True
):
    st.session_state["_simple_checker_df"] = df_source.copy()
    st.session_state["_simple_checker_log"] = []
    st.session_state["_country_review_decisions"] = {}
    st.session_state["_completed_countries"] = []
    st.session_state["_simple_checker_checked"] = False
    st.rerun()
